from django.http import HttpResponse
from django.shortcuts import render , redirect
from django.urls import reverse_lazy
from django.views.generic import ListView, CreateView, UpdateView, DeleteView, View
from django.contrib.auth.models import User
from django.views.defaults import page_not_found

#exportar excel
from django.http.response import HttpResponse
from django.views.generic.base import TemplateView
from openpyxl import Workbook

#importar modelos y fomularios
from cirugia.models import * 
from cirugia.forms import * 
from cirugia.funciones.calculos import *

#para importar libreria de excel
import numpy as np
import xlrd
from copy import copy

import copy as cp

# REPORTE EN EXCEL
class Reporte_proc_excel(TemplateView):
    
    def get(self, request, *args, **kwargs):
        procedimient = procedimiento.objects.all()
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'REPORTE DETALLADO DE LAS CPP CORRESPONDIENTES A ARRIENDOS'
        
        ws['A2']= 'ID'
        ws['B2']= 'Nombre_Procedimiento'
        ws['C2']='Nombre Especialidad'
        
        cont = 3
        for proced in procedimient:
            ws.cell(row=cont, column = 1).value = proced.id
            ws.cell(row=cont, column = 2).value = proced.nombre_proc
            ws.cell(row=cont, column = 3).value = proced.tipo_proc.nombre_tipo_proc
            cont +=1
        
        nombre_reporte = "Reporte_procedimientos_excel.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        content = "attachment; filename = {0}".format(nombre_reporte)
        response['content-Disposition'] = content
        wb.save(response)
        return response

class Reporte_Honorarios_excel(TemplateView):
    
    def get(self, request, *args, **kwargs):
        
        honor = honorario.objects.all()
        wbb = Workbook()
        wss = wbb.active
        wss['A1'] = 'REPORTE DETALLADO DE LOS HONORARIOS'
        
        wss['A2']= 'ID'
        wss['B2']= 'Especialidad'
        wss['C2']= 'Nombre del Procedimiento'
        wss['D2']= 'Canasta'
        wss['E2']= 'Concepto'
        wss['F2']= 'Información'
        wss['G2']= 'Costo no Paramertrizado'
        
        cont = 3
        for i in honor:
            wss.cell(row=cont, column = 1).value = i.id
            wss.cell(row=cont, column = 2).value = format(i.tipo_proc)
            wss.cell(row=cont, column = 3).value = format(i.procedimiento)
            wss.cell(row=cont, column = 4).value = format(i.nombre_canasta)
            wss.cell(row=cont, column = 5).value = format(i.concepto_honorario)
            wss.cell(row=cont, column = 6).value = i.info
            wss.cell(row=cont, column = 7).value = i.costo
            
            cont +=1
        
        nombre_report = "Reporte_honorario_excel.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        content = "attachment; filename = {0}".format(nombre_report)
        response['content-Disposition'] = content
        wbb.save(response)
        return response
    

# pagina de error
def mi_error_404(request):
    nombre_template = '404.html'

    return page_not_found(request, template_name=nombre_template)


class Reporte_Consolidado_excel(TemplateView):
    
    def get(self, request, *args, **kwargs):
        procedimientos  = procedimiento.objects.all()
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'REPORTE CONSOLIDADO DE LOS COSTOS DE LOS PROCEDIMIENTOS QUIRURGICOS'
        
        ws['A2']= 'Nombre_procedimiento'
        ws['B2']= 'canasta'
        ws['C2']= 'valor_canasta'
        ws['D2']= 'desechable'
        ws['E2']= 'medicamento'
        ws['F2']='dispositivos'
        ws['G2']= 'paquete_desechable'
        ws['H2']= 'otro'
        ws['I2']='papeleria'
        ws['J2']= 'duracion'
        ws['K2']= 'consul_anestecia'
        ws['L2']='consul_especialista'
        ws['M2']= 'total_consul'
        ws['N2']= 'sal_instrument'
        ws['O2']='sal_med_ayudante'
        ws['P2']= 'sal_enfermera'
        ws['Q2']= 'sal_total'
        ws['R2']='honor_anestecio'
        ws['S2']= 'dere_sala'
        ws['T2']= 'valor_estancia'
        ws['U2']='dias_estancia'
        ws['V2']= 'valor_estancia_unit'
        ws['W2']= 'honorario_Especialista'
        ws['X2']='total'
        ws['Y2']='ID_del_proc'
        
        #DATOS DE ENTRADA
        name_cansta = nombre_canasta.objects.all()
        honora = honorario.objects.all()
        canast = canasta.objects.all()
        salar = salario.objects.all()
        procedi = procedimiento.objects.all()
        estan = estancia.objects.all()
        tipo_estan = tipo_estancia.objects.all()
        constant = constante.objects.all()
        escalar = 1
        cont = 3
        for i in procedimientos:
            consul = consulta()
            
            temp_tipo_proc = tipo_proc.objects.get(id=i.tipo_proc.id)
            tem_proc = procedimiento.objects.get(id=i.id)
            
            consul.tipo_proc = temp_tipo_proc
            consul.procedimiento = tem_proc
            
            #FUNCION
            datos = calculo(name_cansta,honora, consul, canast, salar, procedi, estan, tipo_estan,constant,escalar)
            datos.resultado()
            
            ws.cell(row=cont, column = 1).value = format(consul.procedimiento)
            ws.cell(row=cont, column = 2).value = datos.resultado()[0].nombre_canasta
            ws.cell(row=cont, column = 3).value = round(datos.resultado()[1])
            ws.cell(row=cont, column = 4).value = round(datos.resultado()[2])
            ws.cell(row=cont, column = 5).value = round(datos.resultado()[3])
            ws.cell(row=cont, column = 6).value = round(datos.resultado()[4])
            ws.cell(row=cont, column = 7).value = round(datos.resultado()[5])
            ws.cell(row=cont, column = 8).value = round(datos.resultado()[6])
            ws.cell(row=cont, column = 9).value = round(datos.resultado()[7])
            ws.cell(row=cont, column = 10).value = round(datos.resultado()[8])
            ws.cell(row=cont, column = 11).value = round(datos.resultado()[9])
            ws.cell(row=cont, column = 12).value = round(datos.resultado()[10])
            ws.cell(row=cont, column = 13).value = round(datos.resultado()[11])
            ws.cell(row=cont, column = 14).value = round(datos.resultado()[12])
            ws.cell(row=cont, column = 15).value = round(datos.resultado()[13])
            ws.cell(row=cont, column = 16).value = round(datos.resultado()[14])
            ws.cell(row=cont, column = 17).value = round(datos.resultado()[15])
            ws.cell(row=cont, column = 18).value =  round(datos.resultado()[16])
            ws.cell(row=cont, column = 19).value = round(datos.resultado()[17])
            ws.cell(row=cont, column = 20).value = round(datos.resultado()[18])
            ws.cell(row=cont, column = 21).value = round(datos.resultado()[19])
            ws.cell(row=cont, column = 22).value = round(datos.resultado()[20])
            ws.cell(row=cont, column = 23).value = datos.resultado()[23]
            ws.cell(row=cont, column = 24).value = datos.resultado()[22]
            ws.cell(row=cont, column = 25).value = format(consul.procedimiento.id)
            cont +=1
        
        nombre_reporte = "Reporte_Consolidado_de_procedimientos.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        content = "attachment; filename = {0}".format(nombre_reporte)
        response['content-Disposition'] = content
        wb.save(response)
        return response

    
#vista de consulta de datos
def consulta_info(request):
    consul = consulta.objects.all()#eso tambien es un dato de entrada como los que estan dentro del IF de abajo
    if request.method == 'POST' :
        form = consultaform(request.POST)
        if form.is_valid(): 
            form.save()
        return redirect('consulta_info')
    else:
        form = consultaform()
    
    if consul.count() != 0:
        consul_ultimo = consul.last()
        
        #DATOS DE ENTRADA
        name_cansta = nombre_canasta.objects.all()
        honora = honorario.objects.all()
        canast = canasta.objects.all()
        salar = salario.objects.all()
        procedi = procedimiento.objects.all()
        estan = estancia.objects.all()
        tipo_estan = tipo_estancia.objects.all()
        constant = constante.objects.all()
        escalar = 0
        #FUNCION
        datos = calculo(name_cansta,honora, consul, canast, salar, procedi, estan, tipo_estan,constant, escalar)
        datos.resultado()
        
        maximo_honorario = datos.resultado()[21]

        print('esta es el objeto ',datos.resultado()[21])
        return render(request,'cirugia/consulta_info.html',{'form':form, 
                                                            'consul_ultimo':consul_ultimo,
                                                            'canasta':datos.resultado()[0].nombre_canasta,
                                                            'valor_canasta':round(datos.resultado()[1]),
                                                            'desechable':round(datos.resultado()[2]),
                                                            'medicamento':round(datos.resultado()[3]),
                                                            'dispositivos':round(datos.resultado()[4]),
                                                            'paquete_desechable':round(datos.resultado()[5]),
                                                            'otro':round(datos.resultado()[6]),
                                                            'papeleria':round(datos.resultado()[7]),
                                                            'duracion':round(datos.resultado()[8]),
                                                            'consul_anestecia':round(datos.resultado()[9]),
                                                            'consul_especialista':round(datos.resultado()[10]),
                                                            'total_consul':round(datos.resultado()[11]),
                                                            'sal_instrument': round(datos.resultado()[12]),
                                                            'sal_med_ayudante': round(datos.resultado()[13]),
                                                            'sal_enfermera': round(datos.resultado()[14]),
                                                            'sal_total': round(datos.resultado()[15]),
                                                            'honor_anestecio': round(datos.resultado()[16]),
                                                            'dere_sala': round(datos.resultado()[17]),
                                                            'valor_estancia':round(datos.resultado()[18]),
                                                            'dias_estancia':round(datos.resultado()[19]),
                                                            'valor_estancia_unit':round(datos.resultado()[20]),
                                                            'honorario':datos.resultado()[21],
                                                            'total':datos.resultado()[22],
                                                            })
    else:
        return render(request,'cirugia/consulta_info.html',{'form':form})

def limpiar_consulta(request):
    consult = consulta.objects.all()
    consult.delete()
    return redirect('consulta_info')

def data_proc_url(request):
    tipo_proc = request.GET.get('tipo_proc')
    procedimiento1 = procedimiento.objects.filter(tipo_proc=tipo_proc).order_by('nombre_proc')
    return render(request,'cirugia/consulta_lista_proc.html',{'procedimientos':procedimiento1})

    
#tipos de procedimietos
def tipo_proc_list(request):
    tipo_pro_1 = tipo_proc.objects.all()
    contexto = {'tipo_procs':tipo_pro_1}
    return render(request, 'cirugia/tipo_proc_list.html',contexto)

def tipo_procEdit(request,id_):
    tipoproc = tipo_proc.objects.get(id = id_)
    if request.method == 'GET':
        form = tipo_procform(instance = tipoproc)
    else:
        form = tipo_procform(request.POST, instance = tipoproc)
        if form.is_valid():
            form.save()
        return redirect('tipo_proc_list')
    contexto = {'form':form}
    return render(request, 'cirugia/tipo_proc_form.html',contexto) 
        
class tipo_procCrear(CreateView):
    model = tipo_proc
    form_class = tipo_procform
    template_name = 'cirugia/tipo_proc_form.html'
    success_url=reverse_lazy('tipo_proc_list')

def tipo_procElim(request, id_):
    tipoproceso = tipo_proc.objects.get(id = id_)
    tipoproceso.delete()
    return redirect('tipo_proc_list')



#procedimientos
def procedimiento_list(request):
    
    # archi = xlrd.open_workbook('C:\\CPP\\entvirtual\\src\\cirugia\\archivo.xlsx', on_demand=True)
    # hoja1 = archi.sheet_by_index(11)
    
    # art = procedimiento.objects.all()
    # contador = 1
    # for i in art:
    #     i.dias_estancia = hoja1.cell_value(contador,2)
    #     i.save()
    #     contador += 1
    
    
    # # ult = tipo_proc.objects.get(id = 6)
    
    # contador = 1
    # for i in art:
    #     if i.id >=56 and i.id <= 76:
    #         name = hoja1.cell_value(contador,0)
        # dur = hoja1.cell_value(contador,1)

    #         # i.tipo_proc = ult           
    #         # i.nombre_proc = name
    
        # i.uvr = dur  
        # i.save()
        # contador += 1
    
    
    
    # archi = xlrd.open_workbook('C:\\CPP\\entvirtual\\src\\cirugia\\archivo.xlsx', on_demand=True)
    # hoja1 = archi.sheet_by_index(10)

    # for i in range(1,hoja1.nrows): 
    #     name = hoja1.cell_value(i,0)
    #     # dur = hoja1.cell_value(i,2)
        
    #     # ult = tipo_proc.objects.get(id=13)
    #     a=procedimiento(
    #                     # tipo_proc = ult,
    #                     nombre_proc = name,
    #                     # duracion_proc = dur,
    #                     )
    #     a.save()

    
    
    procedimiento_1 = procedimiento.objects.all()
    contexto = {'procedimientos':procedimiento_1}
    return render(request, 'cirugia/procedimiento_list.html',contexto)

def procedimientoEdit(request,id_):
    procedimiento1 = procedimiento.objects.get(id = id_)
    if request.method == 'GET':
        form = procedimientoform(instance = procedimiento1)
    else:
        form = procedimientoform(request.POST, instance = procedimiento1)
        if form.is_valid():
            form.save()
        return redirect('procedimiento_list')
    contexto = {'form':form}
    return render(request, 'cirugia/procedimiento_form.html',contexto) 
    
class procedimientoCrear(CreateView):
    model = procedimiento
    form_class = procedimientoform
    template_name = 'cirugia/procedimiento_form.html'
    success_url=reverse_lazy('procedimiento_list')

def procedimientoElim(request, id_):
    procedimiento1 = procedimiento.objects.get(id = id_)
    procedimiento1.delete()
    return redirect('procedimiento_list')




#concepto_honorario
def concepto_honorario_list(request):
    concepto_honorario_1 = concepto_honorario.objects.all()
    contexto = {'concepto_honorarios':concepto_honorario_1}
    return render(request, 'cirugia/concepto_honorario_list.html',contexto)

def concepto_honorarioEdit(request,id_):
    concepto_honorario1 = concepto_honorario.objects.get(id = id_)
    if request.method == 'GET':
        form = concepto_honorarioform(instance = concepto_honorario1)
    else:
        form = concepto_honorarioform(request.POST, instance = concepto_honorario1)
        if form.is_valid():
            form.save()
        return redirect('concepto_honorario_list')
    contexto = {'form':form}
    return render(request, 'cirugia/concepto_honorario_form.html',contexto) 

class concepto_honorarioCrear(CreateView):
    model = concepto_honorario
    form_class = concepto_honorarioform
    template_name = 'cirugia/concepto_honorario_form.html'
    success_url=reverse_lazy('concepto_honorario_list')

def concepto_honorarioElim(request, id_):
    concepto_honorario1 = concepto_honorario.objects.get(id = id_)
    concepto_honorario1.delete()
    return redirect('concepto_honorario_list')



#nombre_canasta
def nombre_canasta_list(request):
    
    # archi = xlrd.open_workbook('C:\\CPP\\entvirtual\\src\\cirugia\\archivo.xlsx', on_demand=True)
    # hoja1 = archi.sheet_by_index(10)

    # for i in range(1,hoja1.nrows): 
    #     name = hoja1.cell_value(i,0)
    #     # dur = hoja1.cell_value(i,2)
        
    #     # ult = tipo_proc.objects.get(id=13)
    #     a=nombre_canasta(
    #                     # tipo_proc = ult,
    #                     nombre_canasta = name,
    #                     # duracion_proc = dur,
    #                     )
    #     a.save()
    
    nombre_canasta_1 = nombre_canasta.objects.all()
    contexto = {'nombre_canastas':nombre_canasta_1}
    return render(request, 'cirugia/nombre_canasta_list.html',contexto)

def nombre_canastaEdit(request,id_):
    nombre_canasta1 = nombre_canasta.objects.get(id = id_)
    if request.method == 'GET':
        form = nombre_canastaform(instance = nombre_canasta1)
    else:
        form = nombre_canastaform(request.POST, instance = nombre_canasta1)
        if form.is_valid():
            form.save()
        return redirect('nombre_canasta_list')
    contexto = {'form':form}
    return render(request, 'cirugia/nombre_canasta_form.html',contexto) 

class nombre_canastaCrear(CreateView):
    model = nombre_canasta
    form_class = nombre_canastaform
    template_name = 'cirugia/nombre_canasta_form.html'
    success_url=reverse_lazy('nombre_canasta_list')

def nombre_canastaElim(request, id_):
    nombre_canasta1 = nombre_canasta.objects.get(id = id_)
    nombre_canasta1.delete()
    return redirect('nombre_canasta_list')



#concepto_canasta
def concepto_canasta_list(request):
    concepto_canasta_1 = concepto_canasta.objects.all()
    contexto = {'concepto_canastas':concepto_canasta_1}
    return render(request, 'cirugia/concepto_canasta_list.html',contexto)

def concepto_canastaEdit(request,id_):
    concepto_canasta1 = concepto_canasta.objects.get(id = id_)
    if request.method == 'GET':
        form = concepto_canastaform(instance = concepto_canasta1)
    else:
        form = concepto_canastaform(request.POST, instance = concepto_canasta1)
        if form.is_valid():
            form.save()
        return redirect('concepto_canasta_list')
    contexto = {'form':form}
    return render(request, 'cirugia/concepto_canasta_form.html',contexto) 

class concepto_canastaCrear(CreateView):
    model = concepto_canasta
    form_class = concepto_canastaform
    template_name = 'cirugia/concepto_canasta_form.html'
    success_url=reverse_lazy('concepto_canasta_list')

def concepto_canastaElim(request, id_):
    concepto_canasta1 = concepto_canasta.objects.get(id = id_)
    concepto_canasta1.delete()
    return redirect('concepto_canasta_list')



#position
def position_list(request):
    position_1 = position.objects.all()
    contexto = {'positions':position_1}
    return render(request, 'cirugia/position_list.html',contexto)

def positionEdit(request,id_):
    position1 = position.objects.get(id = id_)
    if request.method == 'GET':
        form = positionform(instance = position1)
    else:
        form = positionform(request.POST, instance = position1)
        if form.is_valid():
            form.save()
        return redirect('position_list')
    contexto = {'form':form}
    return render(request, 'cirugia/position_form.html',contexto) 

class positionCrear(CreateView):
    model = position
    form_class = positionform
    template_name = 'cirugia/position_form.html'
    success_url=reverse_lazy('position_list')

def positionElim(request, id_):
    position1 = position.objects.get(id = id_)
    position1.delete()
    return redirect('position_list')



#canasta
def canasta_list(request):
    # for i in canasta.objects.all(): 
    #     i.costo_tot = i.costo_und*i.cantidad
    #     i.save()
    
    # archi = xlrd.open_workbook('C:\\CPP\\entvirtual\\src\\cirugia\\archivo.xlsx', on_demand=True)
    # hoja1 = archi.sheet_by_index(10)
    # cont = 1
    # for i in canasta.objects.all(): 
    # #     pos = hoja1.cell_value(cont,2)
    #     if 
    # recuper = canasta.objects.filter(concepto_canasta = 6 )
    # for i in recuper:
    #     i.costo_tot = i.cantidad*i.costo_und
    #     i.save()
    
    canasta_1 = canasta.objects.all()
    contexto = {'canastas':canasta_1}
    return render(request, 'cirugia/canasta_list.html',contexto)

def canastaEdit(request,id_):
    canasta1 = canasta.objects.get(id = id_)
    if request.method == 'GET':
        form = canastaform(instance = canasta1)
    else:
        form = canastaform(request.POST, instance = canasta1)
        if form.is_valid():
            form.save()
            canasta1.costo_tot = canasta1.costo_und*canasta1.cantidad
            canasta1.save()
        return redirect('canasta_list')
    contexto = {'form':form}

    return render(request, 'cirugia/canasta_form.html',contexto) 

def canastaCrear(request):
    if request.method == 'POST' :
        form = canastaform(request.POST)
        if form.is_valid(): 
            form.save()
            canastaa = canasta.objects.last()
            canastaa.costo_tot = canastaa.costo_und*canastaa.cantidad
            canastaa.save()
        return redirect('canasta_list')
    else:
        form = canastaform()
    
    return render(request, 'cirugia/canasta_form.html', {'form':form} )

def canastaElim(request, id_):
    canasta1 = canasta.objects.get(id = id_)
    canasta1.delete()
    return redirect('canasta_list')



#honorario
def honorario_list(request):
    
    archi = xlrd.open_workbook('C:\\CPP\\entvirtual\\src\\cirugia\\archivo.xlsx', on_demand=True)
    hoja1 = archi.sheet_by_index(14)
    
    # cont  = 1
    # for i in honorario.objects.all():
    #     costo = hoja1.cell_value(cont,9)
    #     i.costo = costo
    #     i.save()
    #     cont += 1
        
    #     for m in nombre_canasta.objects.all():
    #         if cana == m.nombre_canasta:
    #             i.nombre_canasta = m
    #             i.save()
        
    #     for n in concepto_honorario.objects.all():
    #         if concep == n.nombre_concep_hon:
    #             i.concepto_honorario = n
    #             i.save()

    #     cont += 1
    
    # cont = 1
    # for i in range(1,hoja1.nrows): 
    #     especi = hoja1.cell_value(i,0)
    #     cir = hoja1.cell_value(i,1)
    #     cana = hoja1.cell_value(cont,2)
    #     concep = hoja1.cell_value(i,3)
    #     info = hoja1.cell_value(i,4)
        

    #     w = honorario()
        
    #     for j in tipo_proc.objects.all():
    #         if especi == j.nombre_tipo_proc:
    #             w.tipo_proc = j
                
    #     for k in procedimiento.objects.all():
    #         if cir == k.nombre_proc:
    #             w.procedimiento=k

    #     for m in canasta.objects.all():
    #         if cana == m.nombre_canasta:
    #             w.canasta = m
                
    #     for n in concepto_honorario.objects.all():
    #         if concep == n.nombre_concep_hon:
    #             w.concepto_honorario = n
                
    #     w.info = info
        
    #     cont += 1
        
    #     w.save()
    
    honorario_1 = honorario.objects.all()
    contexto = {'honorarios':honorario_1}
    return render(request, 'cirugia/honorario_list.html',contexto)

def honorarioEdit(request,id_):
    honorario1 = honorario.objects.get(id = id_)
    if request.method == 'GET':
        form = honorarioform(instance = honorario1)
    else:
        form = honorarioform(request.POST, instance = honorario1)
        if form.is_valid():
            form.save()
        return redirect('honorario_list')
    contexto = {'form':form}
    return render(request, 'cirugia/honorario_form.html',contexto) 

class honorarioCrear(CreateView):
    model = honorario
    form_class = honorarioform
    template_name = 'cirugia/honorario_form.html'
    success_url=reverse_lazy('honorario_list')

def honorarioElim(request, id_):
    honorario1 = honorario.objects.get(id = id_)
    honorario1.delete()
    return redirect('honorario_list')



#constante
def constante_list(request):
    constante_1 = constante.objects.all()
    contexto = {'constantes':constante_1}
    return render(request, 'cirugia/constante_list.html',contexto)

def constanteEdit(request,id_):
    constante1 = constante.objects.get(id = id_)
    if request.method == 'GET':
        form = constanteform(instance = constante1)
    else:
        form = constanteform(request.POST, instance = constante1)
        if form.is_valid():
            form.save()
        return redirect('constante_list')
    contexto = {'form':form}
    return render(request, 'cirugia/constante_form.html',contexto) 

class constanteCrear(CreateView):
    model = constante
    form_class = constanteform
    template_name = 'cirugia/constante_form.html'
    success_url=reverse_lazy('constante_list')

def constanteElim(request, id_):
    constante1 = constante.objects.get(id = id_)
    constante1.delete()
    return redirect('constante_list')



#concepto_salario
def concepto_salario_list(request):
    
    # archi = xlrd.open_workbook('C:\\CPP\\entvirtual\\src\\cirugia\\archivo.xlsx', on_demand=True)
    # hoja1 = archi.sheet_by_index(13)
    

    # for i in range(1,hoja1.nrows): 
    #     especi = hoja1.cell_value(i,0)
        
    #     a= concepto_salario()
    #     a.nombre_concep_sal = especi
    #     a.save()
    
    concepto_salario_1 = concepto_salario.objects.all()
    contexto = {'concepto_salarios':concepto_salario_1}
    return render(request, 'cirugia/concepto_salario_list.html',contexto)

def concepto_salarioEdit(request,id_):
    concepto_salario1 = concepto_salario.objects.get(id = id_)
    if request.method == 'GET':
        form = concepto_salarioform(instance = concepto_salario1)
    else:
        form = concepto_salarioform(request.POST, instance = concepto_salario1)
        if form.is_valid():
            form.save()
        return redirect('concepto_salario_list')
    contexto = {'form':form}
    return render(request, 'cirugia/concepto_salario_form.html',contexto) 

class concepto_salarioCrear(CreateView):
    model = concepto_salario
    form_class = concepto_salarioform
    template_name = 'cirugia/concepto_salario_form.html'
    success_url=reverse_lazy('concepto_salario_list')

def concepto_salarioElim(request, id_):
    concepto_salario1 = concepto_salario.objects.get(id = id_)
    concepto_salario1.delete()
    return redirect('concepto_salario_list')



#tipo_estancia
def tipo_estancia_list(request):
    tipo_estancia_1 = tipo_estancia.objects.all()
    contexto = {'tipo_estancias':tipo_estancia_1}
    return render(request, 'cirugia/tipo_estancia_list.html',contexto)

def tipo_estanciaEdit(request,id_):
    tipo_estancia1 = tipo_estancia.objects.get(id = id_)
    if request.method == 'GET':
        form = tipo_estanciaform(instance = tipo_estancia1)
    else:
        form = tipo_estanciaform(request.POST, instance = tipo_estancia1)
        if form.is_valid():
            form.save()
        return redirect('tipo_estancia_list')
    contexto = {'form':form}
    return render(request, 'cirugia/tipo_estancia_form.html',contexto) 

class tipo_estanciaCrear(CreateView):
    model = tipo_estancia
    form_class = tipo_estanciaform
    template_name = 'cirugia/tipo_estancia_form.html'
    success_url=reverse_lazy('tipo_estancia_list')

def tipo_estanciaElim(request, id_):
    tipo_estancia1 = tipo_estancia.objects.get(id = id_)
    tipo_estancia1.delete()
    return redirect('tipo_estancia_list')


#estancia
def estancia_list(request):
    estancia_1 = estancia.objects.all()
    contexto = {'estancias':estancia_1}
    return render(request, 'cirugia/estancia_list.html',contexto)

def estanciaEdit(request,id_):
    estancia1 = estancia.objects.get(id = id_)
    if request.method == 'GET':
        form = estanciaform(instance = estancia1)
    else:
        form = estanciaform(request.POST, instance = estancia1)
        if form.is_valid():
            form.save()
        return redirect('estancia_list')
    contexto = {'form':form}
    return render(request, 'cirugia/estancia_form.html',contexto) 

class estanciaCrear(CreateView):
    model = estancia
    form_class = estanciaform
    template_name = 'cirugia/estancia_form.html'
    success_url=reverse_lazy('estancia_list')

def estanciaElim(request, id_):
    estancia1 = estancia.objects.get(id = id_)
    estancia1.delete()
    return redirect('estancia_list')



#salario
def salario_list(request):
    
    # archi = xlrd.open_workbook('C:\\CPP\\entvirtual\\src\\cirugia\\archivo.xlsx', on_demand=True)
    # hoja1 = archi.sheet_by_index(13)
    
    # cont = 1
    # for i in salario.objects.all():
    #     i.costo = hoja1.cell_value(cont,5)
    #     i.save()
    #     cont += 1
    
    
    
    # cont = 1
    # for i in salario.objects.all():
        
    #     concep = hoja1.cell_value(cont,3)
    #     ubi = hoja1.cell_value(cont,4)
        
    #     for k in concepto_salario.objects.all():
    #         if concep == k.nombre_concep_sal:
    #             i.concepto_salario = k
        
    #     for n in position.objects.all():
    #         if ubi == n.nombre_act:
    #             i.position = n
        
    #     cont += 1
    #     i.save()
    
    # for i in range(1,hoja1.nrows): 
    #     especi = hoja1.cell_value(i,1)
    #     proc = hoja1.cell_value(i,2)
    #     concep = hoja1.cell_value(i,3)
    #     ubi = hoja1.cell_value(i,4)
        
    #     w = salario()
        
    #     for m in tipo_proc.objects.all():
    #         if especi == m.nombre_tipo_proc:
    #             w.tipo_proc = m 

    #     for j in procedimiento.objects.all():
    #         if proc == j.nombre_proc:
    #             w.procedimiento = j
                
    #     for k in concepto_salario.objects.all():
    #         if concep == k.nombre_concep_sal:
    #             w.concepto_salario = k
        
    #     for n in position.objects.all():
    #         if prubioc == n.nombre_act:
    #             w.position = n
                
    #     w.save()
        
    
    salario_1 = salario.objects.all()
    contexto = {'salarios':salario_1}
    return render(request, 'cirugia/salario_list.html',contexto)

def salarioEdit(request,id_):
    salario1 = salario.objects.get(id = id_)
    if request.method == 'GET':
        form = salarioform(instance = salario1)
    else:
        form = salarioform(request.POST, instance = salario1)
        if form.is_valid():
            form.save()
        return redirect('salario_list')
    contexto = {'form':form}
    return render(request, 'cirugia/salario_form.html',contexto) 

class salarioCrear(CreateView):
    model = salario
    form_class = salarioform
    template_name = 'cirugia/salario_form.html'
    success_url=reverse_lazy('salario_list')

def salarioElim(request, id_):
    salario1 = salario.objects.get(id = id_)
    salario1.delete()
    return redirect('salario_list')



#tiempo_proc
# def tiempo_proc_list(request):
#     tiempo_proc_1 = tiempo_proc.objects.all()
#     contexto = {'tiempo_procs':tiempo_proc_1}
#     return render(request, 'cirugia/tiempo_proc_list.html',contexto)

# def tiempo_procEdit(request,id_):
#     tiempo_proc1 = tiempo_proc.objects.get(id = id_)
#     if request.method == 'GET':
#         form = tiempo_procform(instance = tiempo_proc1)
#     else:
#         form = tiempo_procform(request.POST, instance = tiempo_proc1)
#         if form.is_valid():
#             form.save()
#         return redirect('tiempo_proc_list')
#     contexto = {'form':form}
#     return render(request, 'cirugia/tiempo_proc_form.html',contexto) 

# class tiempo_procCrear(CreateView):
#     model = tiempo_proc
#     form_class = tiempo_procform
#     template_name = 'cirugia/tiempo_proc_form.html'
#     success_url=reverse_lazy('tiempo_proc_list')

# def tiempo_procElim(request, id_):
#     tiempo_proc1 = tiempo_proc.objects.get(id = id_)
#     tiempo_proc1.delete()
#     return redirect('tiempo_proc_list')