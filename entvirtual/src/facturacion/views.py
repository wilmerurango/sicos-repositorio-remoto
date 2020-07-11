from django.http import HttpResponse
from django.shortcuts import render , redirect
from django.db.models import Subquery
from django.urls import reverse_lazy
from django.views.generic import ListView, CreateView, UpdateView, DeleteView, View

#importar funciones creadas
from facturacion.funciones.calculater import *
from facturacion.funciones.defprov import *
from facturacion.funciones.defserv import *

from facturacion.models import * 
from facturacion.forms import * 
from django.http import FileResponse
from django.conf import settings
import io 
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.platypus.tables import Table
from reportlab.platypus import SimpleDocTemplate
import time
import copy
import numpy as np
from django.urls import path

#login
from django.contrib.auth.models import User
from django.contrib.auth.forms import UserCreationForm
# from anaconda_navigator.config import user
# from django.contrib.auth import user
from django.contrib.auth.models import User
import getpass


#exportar excel
from openpyxl import Workbook
from django.http.response import HttpResponse
from django.views.generic.base import TemplateView

#==================================================================================================
#|||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||
#==================================================================================================
#==========******CREACION DE USUARIO******==========================================================
#==================================================================================================
#|||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||
#==================================================================================================

class RegistroUsuario(CreateView):
    model = User
    template_name = "Entrada/register.html"
    form_class = RegistroForm
    success_url = reverse_lazy('home')   
    
#cambiar contraseña


#==================================================================================================
#|||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||
#==================================================================================================
#==========******REPORTES EN EXCEL******==========================================================
#==================================================================================================
#|||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||
#==================================================================================================

    
class Reporte_arriendo_excel(TemplateView):
    def get(self, request, *args, **kwargs):
        arriendos = cpp_arriendo_detalle.objects.all()
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'REPORTE DETALLADO DE LAS CPP CORRESPONDIENTES A ARRIENDOS'
        
        ws['A2']= 'ID'
        ws['B2']= 'N° Factura'
        ws['C2'] = 'Nombre Empresa'
        ws['D2'] = 'Cuenta Retenedora'
        ws['E2'] = 'Centro de Costo'
        ws['F2'] = 'Inductor'
        ws['G2'] = 'Retención'
        ws['H2'] = 'Cuenta C. Costo'
        ws['I2'] = 'Cuenta Especifica'
        ws['J2'] = 'Valor'
        ws['K2'] = 'Fecha'
        
        cont = 3
        for arriendo in arriendos:
            ws.cell(row=cont, column = 1).value = arriendo.id
            ws.cell(row=cont, column = 2).value = arriendo.cpp_arriendo
            ws.cell(row=cont, column = 3).value = arriendo.name_arri
            ws.cell(row=cont, column = 4).value = arriendo.cuenta_reten
            ws.cell(row=cont, column = 5).value = arriendo.centro_costo
            ws.cell(row=cont, column = 6).value = arriendo.inductor_arri
            ws.cell(row=cont, column = 7).value = arriendo.reten
            ws.cell(row=cont, column = 8).value = arriendo.num_cuenta
            ws.cell(row=cont, column = 9).value = arriendo.cuenta_especific
            ws.cell(row=cont, column = 10).value = arriendo.valor_cpp_arri_detal
            ws.cell(row=cont, column = 11).value = arriendo.fecha_cpp_arri_detal
            cont +=1
        
        nombre_reporte = "Reporte_arriendo_excel.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        content = "attachment; filename = {0}".format(nombre_reporte)
        response['content-Disposition'] = content
        wb.save(response)
        return response
                
class Reporte_especialista_excel(TemplateView):
    def get(self, request, *args, **kwargs):
        especialistas = fac_especialista_detalle.objects.all()
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'REPORTE DETALLADO DE LAS CPP CORRESPONDIENTES A ESPECIALISTAS'
        
        ws['A2']= 'ID'
        ws['B2']= 'Numero_factura'
        ws['C2'] = 'centro_costo'
        ws['D2'] = 'actividad'
        ws['E2'] = 'tipo_fact'
        ws['F2'] = 'valor'
        ws['G2'] = 'fecha_facturacion'

        
        cont = 3
        for especialista in especialistas:
            ws.cell(row=cont, column = 1).value = especialista.id
            ws.cell(row=cont, column = 2).value = especialista.fac_especialista.id
            ws.cell(row=cont, column = 3).value = especialista.centro_costo.name_ccos
            ws.cell(row=cont, column = 4).value = especialista.centro_actividad.actividad.name_act
            ws.cell(row=cont, column = 5).value = especialista.tipo_fact.name_fact
            ws.cell(row=cont, column = 6).value = especialista.valor
            ws.cell(row=cont, column = 7).value = especialista.fechafac_detalle
            
            cont +=1
        
        nombre_reporte = "Reporte_Escpecialistas_excel.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        content = "attachment; filename = {0}".format(nombre_reporte)
        response['content-Disposition'] = content
        wb.save(response)
        return response  
            
#==================================================================================================
#|||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||
#==================================================================================================
#==========******REPORTES EN PDF******=============================================================
#==================================================================================================
#|||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||
#==================================================================================================
#Reporte de Arriendos
def Reporte_CppArriendo_Pdf(request, id_):
    response = HttpResponse(content_type='application/pdf')
    buffer=io.BytesIO()
    archivo_imagen = settings.STATICFILES_URL + '/img/clinicario_log.jpeg'
    p=canvas.Canvas(buffer, pagesize=A4)
    
    #ENCABEZADO ==============================================================================
    y1=700# ESTE ES EL PUNTO D EREFRENCIA DEL ENCABEZADO Y DE TODO EL DOUMENTO
    p.drawImage(archivo_imagen,40,y1-25, 120,90, preserveAspectRatio=True)
    p.drawString(210, y1+45, "FUNDACIÓN CLÍNICA DEL RÍO")
    p.drawString(240, y1+30, "NIT. 900 540 156-1")
    p.drawString(230, y1, "CUENTA POR PAGAR")
    p.drawString(460, y1+45, "Código:")
    p.drawString(505, y1+45, "xxxx")
    p.drawString(460, y1+30, "Version:")
    p.drawString(460, y1+15, "Emision:")
    p.drawString(460, y1, "Fecha:")
    p.drawString(505, y1, time.strftime("%d/%m/%y"))
    #FIN ENCABEZADO ==============================================================================

    
    #ESTE ES EL CUERPO DEL PDF=====================================================================
    y = y1-55 #punto de referencia para mover todos los objetos
    
    obj1=cpp_arriendo.objects.get(id=id_)
    obj2=inductor_arri.objects.filter(arriendo=obj1.arriendo)
    obj3=cpp_arriendo_detalle.objects.filter(cpp_arriendo=id_)
    cuenta_aux_arri = cuenta_arriendo_aux.objects.filter(arriendo = obj1.arriendo.id)
    contar_obj3=obj3.count()

    inductor_arriendo_cpp = factura_pdf(obj1,obj2)
    vector_resultado = inductor_arriendo_cpp.calculo_cpp_arriendo() 
    sumaT = inductor_arriendo_cpp.suma_cpp_arriendo() 

    p.line(40,y+11, 540,y+11)
    p.drawString(40, y, "N° Factura: "), p.drawString(105, y, str(obj1.id)), p.drawString(350, y, "Valor de la Factura:"), p.drawString(460, y, '$' + str('{:,}'.format(obj1.valor_cpp_arri)))
    p.drawString(350, y-15, "Valor a Cancelar:"), p.drawString(460, y-15, '$' + str('{:,}'.format(round(obj1.valor_cpp_arri-(obj1.valor_cpp_arri*obj1.reten/100)))))
    p.drawString(40, y-15, "Empresa:"), p.drawString(105, y-15, obj1.arriendo.name_arri)
    p.drawString(40, y-30, "NIT:"), p.drawString(105, y-30,obj1.arriendo.id_arri)
    p.line(40,y-32, 540,y-32)
    
    p.drawString(185, y-63, "DETALLE DE LA CUENTA POR PAGAR")
    p.line(40,y-66, 540,y-66)
    
    
    p.drawString(40, y-80, "Centro de Costo"),  p.drawString(210, y-80, "Inductor"),  p.drawString(320, y-80, "N° Cuenta"), p.drawString(440, y-80, "Costo asignado")
    p.drawString(40, y-81, "_______________"), p.drawString(210, y-81,  "________"), p.drawString(320, y-81,  "_________"), p.drawString(440, y-81,  "______________")
    
    # ,p.drawString(260, y-120, "Retención")
    # ,p.drawString(260, y-121,  "_________")
    
    # esta condicion es para ubicar la distrinucion de los costes en los distintos centros de coste
    if contar_obj3 == 0:# el cuerpo se llena con los datos registrados enlazados a inductor_arriendo
        j=0
        for i in obj2:
            a = y-95-(j*15)
            p.drawString(40,a, i.centro_costo.name_ccos)
            p.drawString(210, a, str(i.induc)+"%")
            p.drawString(320, a, str(i.cuenta_especific))
            p.drawString(440, a, "$"+str('{:,}'.format(round(vector_resultado[j,0]))))
            j += 1
        
        p.line(40,a-3, 550, a-3)
        interl = 1   
        for s in cuenta_aux_arri:
            if s.cuenta[0:4] == '2365':
                p.drawString(40,a-(interl*15), s.name_cuenta)
                p.drawString(210, a-(interl*15),  str(obj1.reten)+"%")
                p.drawString(320, a-(interl*15),  s.cuenta)
                p.drawString(440, a-(interl*15),  str(round(obj1.valor_cpp_arri*obj1.reten/100)))
            else:
                p.drawString(40,a-(interl*15), s.name_cuenta)
                p.drawString(210, a-(interl*15),  "-")
                p.drawString(320, a-(interl*15),  s.cuenta)
                p.drawString(440, a-(interl*15),  str('{:,}'.format(round((obj1.valor_cpp_arri)-(obj1.valor_cpp_arri*obj1.reten/100)))))
                
            interl +=1
    else:# el cuerpo se llena con los registros encontrados en el cpp_arriendo_detalle
        a = y-95
        j=0
        for i in obj3:
            #dibujar un alindea divisoria
            if i.cuenta_especific[0:4] == "2365":
                p.line(40,a-j*15,550, a-j*15)
                a= a-15
                
            
            p.drawString(40,a-j*15, i.centro_costo)
            p.drawString(210, a-j*15, str(i.inductor_arri)+"%")
            p.drawString(320, a-j*15, str(i.cuenta_especific))
            p.drawString(440, a-j*15, "$"+str('{:,}'.format(round(i.valor_cpp_arri_detal))))
            j += 1
    
    
        
        
    #PIE DE PAGINA =================================================================================================
    p.drawString(40,100, "_____________________________________________________________________________")
    p.drawString(170,80, "CUENTA POR PAGAR GENERADA POR SICOS")
    p.drawString(235,65 , "Fundación Clínia del Río")
    p.drawString(161,50, "Dirección: Cra. 3 #128, Montería, Córdoba, Colombia")
    p.drawString(228,35, "Telefono: +57 311 7623443")
    
    p.setFont('Helvetica', 7) 
    # Captura usuario actual del Equipo
    current_user = request.user
    if User.is_active: 
        usuarioEquipo = getpass.getuser()  
        p.drawString(100,20, "Esta Factura fue Realizada en el Equipo " + '"'+str(usuarioEquipo)+'"'+", Por el Usuario " + '"' + str(current_user) + '"'+ ' el dia '+ time.strftime("%a-%d/%m/%y")+' a las '+ time.strftime(" %r"))
    
    
    #FIN PIE DE PAGINA =================================================================================================

    p.showPage()
    p.save()
    # buffer.seek(0)
    p=buffer.getvalue()
    buffer.close()
    response.write(p)
    return response

#Reporte de Especialistas
def Reporte_CppEspecialista_Pdf(request, id_):
    response = HttpResponse(content_type='application/pdf')
    buffer=io.BytesIO()
    archivo_imagen = settings.STATICFILES_URL + '/img/clinicario_log.jpeg'
    archivo_imagen_logo = settings.STATICFILES_URL + '/img/sicos_logo_trans.jpg'
    p=canvas.Canvas(buffer, pagesize=A4)
    
    p.setFont('Helvetica', 7) 
    
    #ENCABEZADO ==============================================================================
    y1=750# ESTE ES EL PUNTO DE REFRENCIA DEL ENCABEZADO Y DE TODO EL DOUMENTO
    p.drawImage(archivo_imagen,37,y1-8, 95,65, preserveAspectRatio=True)
    p.drawImage(archivo_imagen_logo,485,16, 80,50, preserveAspectRatio=True)
    p.drawString(210, y1+45, "FUNDACIÓN CLÍNICA DEL RÍO")
    p.drawString(240, y1+35, "NIT. 900 540 156-1")
    p.drawString(160, y1+20, "DISTRIBUCIÓN DE COSTOS Y GASTOS POR PAGAR POR ")
    p.drawString(160, y1+10, "SERVICIOS MÉDICOS POR EVENTO PERSONA NATURAL ")
    
    p.drawString(460, y1+45, "Código:"), p.drawString(490, y1+45, "COS-FO-01")
    p.drawString(460, y1+35, "Version:"), p.drawString(490, y1+35, "01")
    p.drawString(460, y1+25, "Emision:"), p.drawString(490, y1+25, "07-07/2020")
    p.drawString(460, y1+15, "Página:"), p.drawString(490, y1+15, "1 de 1")
    
    #FIN ENCABEZADO ==============================================================================

    #datos
    y = y1-30 #punto de referencia para mover todos los objetos
    
    obj2 = fac_especialista.objects.get(id=id_)
    obj1 = contrato.objects.get(especialista = obj2.especialista)
    obj3 = uvt.objects.get(tarifa = obj2.tarifa.id)
    obj4 = reten_383.objects.filter(tarifa = obj2.tarifa.id)
    
    aaa = fac_especialista_detalle.objects.filter(fac_especialista=id_)
    bbb = tipo_fact.objects.all()
    ccc = centro_actividad.objects.all()
    ddd = contrato.objects.get(especialista = obj2.especialista.id)
    # reten_provi = cuenta_reten.objects.get(contrato=ddd.id)
    
    suma_parcial=factura_pdf(aaa, bbb, ccc, ddd)
    (resul,total_sum,product, vec_ev, vec_mf, unicoe, unico, vec_eva, unicoea)=suma_parcial.opera_especialistas()
    
    cont_ev = len(vec_ev)
    cont_mf = len(vec_mf)
    cen_act =centro_actividad.objects.all()
    

    #CUERPO DEL PDF
    p.drawString(250, y+27, "Fecha Digitación:"), p.drawString(310, y+27, time.strftime("%d/%m/%y"))
    
    p.line(40,y+25, 550,y+25)
    p.drawString(40, y+17, "N° Factura: "), p.drawString(95, y+17, str(id_))
    p.drawString(40, y+10, "Especialista:"), p.drawString(95, y+10, obj2.especialista.name_esp + ' ' + obj2.especialista.apellidos_esp)
    p.drawString(40, y+3, "Cédula:"), p.drawString(95, y+3,obj2.especialista.id_esp)

    p.drawString(310, y+17, "Total Facturado: "), p.drawString(460,y+17, "$ " + str('{:,}'.format(round(float(sum(resul)-sum(vec_mf)+obj1.valor)))))
    p.drawString(310, y+10, "Glosa del mes: "), p.drawString(460, y+10,  "$ " + str('{:,}'.format(obj2.glosa)))
    p.drawString(310, y+3, "Base Liquidación Salud o Pensión: "), p.drawString(460, y+3,  "$ " + str('{:,}'.format(round(float(sum(resul)-sum(vec_mf)+obj1.valor-obj2.glosa-sum(vec_eva))))))
    p.line(40,y, 550,y)
    
    p.setFont('Helvetica-Bold', 7)
    p.drawString(190, y-10, "LIQUIDACIÓN DE RETENCIÓN EN LA FUENTE - HONORARIOS")
    p.setFont('Helvetica', 7)
    
    p.line(40,y-15, 180,y-15)
    p.drawString(40,y-23, "S. Social Reglamentaria:"), p.drawString(160,y-23, obj1.razon_social_reglament)
    p.drawString(40,y-31, "Retención DEL 11%:"), p.drawString(160,y-31, obj1.reten_11)
    p.drawString(40,y-39, "Aplica Retención por Art. 383:"), p.drawString(160,y-39, obj1.reten_art_383)
    p.drawString(40,y-47, "Dependiente a Cargo:"), p.drawString(160,y-47, obj1.dependiente_cargo)
    p.drawString(40,y-55, "Retención Tarifa General  del 10%:"), p.drawString(160,y-55, obj1.reten_10)
    p.drawString(40,y-63, "Obligado a Cotizar Pensión:"), p.drawString(160,y-63, obj1.pension_obligado)
    p.line(40,y-65, 180,y-65)
    
    p.line(40,y-72, 180,y-72)
    p.drawString(40,y-80, "Valor UVT:"), p.drawString(80,y-80, '$ ' + str('{:,}'.format(obj3.valor_uvt)))
    p.line(40,y-82, 180,y-82)
    
    p.line(200,y-15, 550,y-15)
    p.drawString(320,y-23, "Tabla de Retención Art. 383 ET.")
    p.line(200,y-25, 550,y-25)
    
    p.setFont('Helvetica', 6)
    
    #llenar tabla de retencion articulo 383
    contw = 0
    for t in obj4:
        if t.id == obj4.first().id:
            p.drawString(205,y-32-(contw*8), "Mayor de " + str(int(t.minimo))+" UVT "), p.drawString(305,y-32-(contw*8), "Hasta "+str(int(t.maximo)) + " UVT "), p.drawString(380,y-32-(contw*8), str(int(t.porcent))+"%" )
        else:
            p.drawString(205,y-32-(contw*8), "Mayor de " + str(int(t.minimo))+" UVT "), p.drawString(305,y-32-(contw*8), "Hasta "+str(int(t.maximo)) + " UVT "), p.drawString(380,y-32-(contw*8), "(Ingreso Gravado en UVT - " + str(int(t.resta))+ ") * "+ str(round(float(t.porcent)))+"% + " + str(int(t.adicion)))
        contw += 1
    
    #********************************************
    calculo = factura_pdf(obj1, obj2, obj3, obj4)
    
    (honorario_cal,
     incr_aport_pension_cal, 
     incr_solida_pensional_cal,
     incr_aport_salud_cal, 
     incr_aport_arl_cal, 
     incr_aport_vol_pension_cal,
     aport_volun_empleador_Cal,
     indemni_lab_cal, 
     re_rent_exent_lab_cal, 
     re_deduc_rent_exent_cal, 
     re_tope_rent_exent_lab_cal, 
     re_total_base_grav_reten_cal,
     re_valor_reten_cal,
     deduc_int_prest_vivienda_cal, 
     deduc_plan_comp_salud_cal, 
     deduc_depen_cargo_cal, 
     re_base_grav_reten_uvt_cal, 
     re_fuente_uvt_cal,
     aport_cuenta_afc_cal,
     
     total_ingre_no_rent_cal,
     total_deducciones_cal,
     total_rent_exten_dos_cal
     
     ) = calculo.retencion_esp()
    #********************************************
    
    #ingresos que no coonstituyen renta
    p.line(40,y-92, 290,y-92)
    p.setFont('Helvetica-Bold', 6)
    p.drawString(120,y-100, "DEPURACIÓN DE RETENCIÓN")
    p.setFont('Helvetica', 6)
    p.line(40,y-102, 290,y-102)
    p.drawString(40,y-110, "Ingresos no Constitutivos de Renta")
    p.line(40,y-111, 133,y-111)
    p.drawString(40,y-118, "Art. 55 ET-Aportes a Pension 16%"), p.drawString(250,y-118, "$ " + str('{:,}'.format(incr_aport_pension_cal) ))
    p.drawString(40,y-126, "Fondo de Solidaridad Pensional 1%"), p.drawString(250,y-126, "$ " + str('{:,}'.format(incr_solida_pensional_cal) ))
    p.drawString(40,y-134, "Art. 56 ET-Aportes a Salud 12.5%"), p.drawString(250,y-134, "$ " + str('{:,}'.format(incr_aport_salud_cal)))
    p.drawString(40,y-142, "Aportes ARL"), p.drawString(250,y-142, "$ " + str('{:,}'.format(incr_aport_arl_cal)))
    p.drawString(40,y-150, "Aportes Voluntarios a Pension (< 25% / Salario)"),p.drawString(190,y-150, "$ " + str('{:,}'.format(obj2.aport_volun_pension))), p.drawString(250,y-150, "$ " + str('{:,}'.format(incr_aport_vol_pension_cal)))
    p.line(250,y-152, 290,y-152)
    p.drawString(40,y-158, "TOTAL INGRESOS NCR"), p.drawString(250,y-158, "$ " + str('{:,}'.format(total_ingre_no_rent_cal)))
    
    #deducciones
    p.drawString(40,y-166, "Deducciones")
    p.line(40,y-168, 75,y-168)
    p.drawString(40,y-176, "Intereses por Prestamo de Vivienda hasta 100 UVT"), p.drawString(190,y-176, "$ " + str('{:,}'.format(obj2.int_pre_vivi))), p.drawString(250,y-176, "$ " + str('{:,}'.format(deduc_int_prest_vivienda_cal)))
    p.drawString(40,y-184, "Art. 367 ET-Plan Complementario de Salud"), p.drawString(190,y-184, "$ " + str('{:,}'.format(obj2.plan_comp_salud))), p.drawString(250,y-184, "$ " + str('{:,}'.format(deduc_plan_comp_salud_cal)))
    p.drawString(40,y-192, "Art. 367 ET-Dependendiente a Cargo"),  p.drawString(250,y-192, "$ " + str('{:,}'.format(deduc_depen_cargo_cal)))
    p.line(250,y-194, 290,y-194)
    p.drawString(40,y-200, "TOTAL DEDUCCIONES"), p.drawString(250,y-200, "$ " + str('{:,}'.format(total_deducciones_cal)))

    #rentas extensas
    p.drawString(40,y-208, "Rentas Extensas")
    p.line(40,y-210, 93,y-210)
    p.drawString(40,y-218, "Art. 126-4 Aportes Cuentas AFC"),p.drawString(190,y-218, "$ " + str('{:,}'.format(obj2.aport_afc))),p.drawString(250,y-218, "$ " + str('{:,}'.format(aport_cuenta_afc_cal)))
    p.drawString(40,y-226, "Art. 126-1 Aportes Voluntarios Empleador"),p.drawString(190,y-226, "$ " + str('{:,}'.format(obj2.aport_volun_emple))), p.drawString(250,y-226, "$ " + str('{:,}'.format(aport_volun_empleador_Cal)))
    p.drawString(40,y-234, "Indemnizaciones Laborales"), p.drawString(190,y-234, "$ " + str(obj2.indem_lab)), p.drawString(250,y-234, "$ " + str(indemni_lab_cal))
    p.line(250,y-236, 290,y-236)
    p.drawString(40,y-242, "TOTAL RENTA EXTENSA"), p.drawString(250,y-242, "$ " + str('{:,}'.format(total_rent_exten_dos_cal)))

    #resto
    p.drawString(40,y-252, "Renta Extensa Laboral"), p.drawString(250,y-252, "$ " + str('{:,}'.format(re_rent_exent_lab_cal)))
    p.drawString(40,y-260, "Total Deducciones + Rentas Extensas"),  p.drawString(250,y-260, "$ " + str('{:,}'.format(re_deduc_rent_exent_cal)))
    p.drawString(40,y-268, "Tope Deducciones + Rentas Extensas"),  p.drawString(250,y-268, "$ " + str('{:,}'.format(re_tope_rent_exent_lab_cal)))
    p.drawString(40,y-276, "Total Base Gravable para Retención"),  p.drawString(250,y-276, "$ " + str('{:,}'.format(re_total_base_grav_reten_cal)))
    p.drawString(40,y-284, "Base Gravable expresada en UVT"),  p.drawString(250,y-284, "$ " + str('{:,}'.format(re_base_grav_reten_uvt_cal)))
    p.drawString(40,y-292, "Rentención en la fuente expresada en UVT"), p.drawString(250,y-292, "$ " + str('{:,}'.format(re_fuente_uvt_cal)))
    p.drawString(40,y-300, "VALOR RETENCIÓN"), p.drawString(250,y-300, "$ " + str('{:,}'.format(re_valor_reten_cal)))
    p.line(40,y-302, 290,y-302)
    
    
    #DETALLE DE CPP ESPECIALISTAS
    p.setFont('Helvetica', 6)
    
    contador = 1
    contador_2 = 1
    contador_3 = 0
    for m in bbb:
        mf = aaa.filter(tipo_fact=m.id)
        
        base_alt = 320 + contador_3*20
        

        #calcular numero de cuentas repetidas 
        uni_cent_costo = []
        for i in mf:
            if i.centro_costo.name_ccos not in uni_cent_costo:
                uni_cent_costo.append(i.centro_costo.name_ccos)
        
        cont_uni_cent_costo = len(uni_cent_costo) 
        
        if cont_uni_cent_costo != 0:
            p.setFont('Helvetica-Bold', 6)
            p.drawString(110,y-base_alt-contador_2*8+10, "DETALLE CPP DEVENGADA POR : " + m.name_fact)
            p.setFont('Helvetica', 6)
            p.line(40,y-base_alt-contador_2*8+10-2, 360,y-base_alt-contador_2*8+10-2)

        
        for i in range(cont_uni_cent_costo):
           
            p.drawString(40,y-base_alt-(contador_2)*8, uni_cent_costo[i])
            p.line(40,y-base_alt-(contador_2)*8 -2, 100,y-base_alt-(contador_2)*8 -2)
            
            contador = contador_2
            for j in mf:
                
                if j.centro_costo.name_ccos == uni_cent_costo[i]:
                    contador += 1
                    p.drawString(40,y-base_alt-(contador*8), j.centro_actividad.actividad.name_act)
                    
                    p.drawString(220,y-base_alt-(contador*8), "$ " + str('{:,}'.format(j.centro_actividad.actividad.valor_iss)))
                    
                    if contador_3 == 0 :
                        p.drawString(190,y-base_alt-(contador*8), str(int(j.valor)))
                        p.drawString(270,y-base_alt-(contador*8), "$ " + str('{:,}'.format(j.centro_actividad.actividad.valor_iss*(1+(j.centro_actividad.actividad.pct_iss)/100))))
                        p.drawString(320,y-base_alt-(contador*8), "$ " + str('{:,}'.format((j.centro_actividad.actividad.valor_iss*(1+(j.centro_actividad.actividad.pct_iss)/100))*j.valor)))
                    else:
                        p.drawString(320,y-base_alt-(contador*8),"$ " + str('{:,}'.format(j.valor)))
                                
            contador_2 = contador + 1
            
        if cont_uni_cent_costo != 0:    
            p.line(40,y-base_alt-(contador*8)-2, 360,y-base_alt-(contador*8)-2)

        contador_3 += 1

    p.line(305,y-92, 550,y-92)
    p.setFont('Helvetica-Bold', 6)
    p.drawString(345,y-100, "DISTRIBUCIÓN CONTABLE DEL COSTO/GASTO")
    p.setFont('Helvetica', 6)
    p.line(305,y-102, 550,y-102)
    
    p.drawString(305,y-112, "Monto Fijo"), p.drawString(405,y-112, "Productividad")
    p.line(305,y-114, 363,y-114), p.line(405,y-114, 448,y-114)
    p.drawString(305,y-122, "$ " + str('{:,}'.format(obj1.valor))),p.drawString(405,y-122, "$ " + str('{:,}'.format(round(float((sum(vec_mf))))))) 
    
    p.line(305,y-124, 550,y-124)
    
    p.drawString(305,y-140, "Cuenta"), p.drawString(365,y-140, "Nombre Cuenta"), p.drawString(470,y-140, "%"), p.drawString(500,y-140, "Valor")
    p.line(305,y-142, 328,y-142), p.line(365,y-142, 415,y-142), p.line(470,y-142, 484,y-142),  p.line(500,y-142, 550,y-142)
    
    
    #colocar cuentas y valores de tipo monto fijo
    cont_unico =len(unico)
    base_alt = 152 
    contador = 0  
    for i in range(cont_unico):
        
        p.drawString(305,y-base_alt-i*10,str(unico[i])), p.drawString(470,y-base_alt-i*10, str(round(float((vec_mf[i]/sum(vec_mf))*100),1)) + " %" ), p.drawString(500,y-base_alt-i*10, "$ " + str('{:,}'.format(round(float((vec_mf[i]/sum(vec_mf))*obj1.valor)))))
        
        contt = 0
        for j in ccc:
            if unico[i]==j.cuenta and contt==0:
                p.drawString(365,y-base_alt-i*10,j.nombre_centro_act)
                contt += 1
                
        contador += 1  
         
        if i == cont_unico-1:
            p.line(500,y-base_alt-i*10-2, 550, y-base_alt-i*10-2)
            p.drawString(441,y-base_alt-i*10-10, "Total Monto Fijo: "), p.drawString(500,y-base_alt-i*10-10, "$ "+ str('{:,}'.format(round(obj1.valor))))
    
    
    #colocar cuenytas y valores de tipo evento
    cont_unicoe = len(unicoe)
    contadorr=0
    for j in range(cont_unicoe):
        p.drawString(305,y-base_alt-j*10-((contador + 1.5)*10),str(unicoe[j])), p.drawString(500,y-base_alt-j*10-((contador + 1.5)*10),"$ " + str('{:,}'.format(round(float(vec_ev[j])))))
        
        contt = 0
        for k in ccc:
            if unicoe[j]==k.cuenta and contt==0:
                p.drawString(365,y-base_alt-j*10-((contador + 1.5)*10), k.nombre_centro_act)
                contt += 1
                
        if j == cont_unicoe-1:
            p.line(500,y-base_alt-j*10-((contador+ 1.5)*10) -2, 550, y-base_alt-j*10-((contador + 1.5)*10 )-2)
            p.drawString(450,y-base_alt-j*10-((contador + 1.5)*10) - 10, "Total Evento: ")
            p.drawString(500,y-base_alt-j*10-((contador + 1.5)*10) - 10, "$ "+ str('{:,}'.format(round(float(sum(vec_ev))))))

        contadorr += 1
        
        
    #colocar cuentas y valores de tipo evento-arriendo
    cont_unicoea = len(unicoea)
    contadorr += 1
    for j in range(cont_unicoea):
        p.drawString(305,y-base_alt-j*10-((contadorr + 1.5)*10),str(unicoea[j])), p.drawString(500,y-base_alt-j*10-((contadorr + 1.5)*10),"$ " + str('{:,}'.format(round(float(vec_eva[j])))))
        
        contt = 0
        for k in ccc:
            if unicoea[j]==k.cuenta and contt==0:
                p.drawString(365,y-base_alt-j*10-((contadorr + 1.5)*10), k.nombre_centro_act)
                contt += 1
                
        if j == cont_unicoea-1:
            p.line(500,y-base_alt-j*10-((contadorr+ 1.5)*10) -2, 550, y-base_alt-j*10-((contadorr + 1.5)*10 )-2)
            p.drawString(450,y-base_alt-j*10-((contadorr + 1.5)*10) - 10, "Total Arriendo: ")
            p.drawString(500,y-base_alt-j*10-((contadorr + 1.5)*10) - 10, "$ "+ str('{:,}'.format(round(float(sum(vec_eva))))))

        contadorr += 1
        
    p.setFont('Helvetica-Bold', 6)
    p.drawString(433,y-base_alt-(contadorr-1)*10-((contador + 4)*10), "TOTAL FACTURA: " )       
    p.drawString(500,y-base_alt-(contadorr-1)*10-((contador + 4)*10), "$ " + str('{:,}'.format(round(float(sum(resul)-sum(vec_mf)+obj1.valor)))))   
        
    p.setFont('Helvetica', 6)
    
    if cont_unico==0:#Verificar que la variable contt este inicializada, 
        contt =1
        
         
    ccont =0 #cuenta cuantas veces se entro en los if, ademas sirve para interlineado
    acum = 0    #acumulador de retenciones      
    if ddd.reten_10 =='SI':#calcular y ubicar la retencion del 10 %
        
        p.drawString(305,y-base_alt-(contadorr-1)*10-((contador + 4 + contt + ccont)*10), "23651501")
        p.drawString(365,y-base_alt-(contadorr-1)*10-((contador + 4 + contt + ccont)*10), "Retención 10%")
        p.drawString(500,y-base_alt-(contadorr-1)*10-((contador + 4 + contt + ccont)*10),"$ "+str('{:,}'.format(round(float((sum(vec_ev)+obj1.valor)*0.10)))))
        ccont += 1 
        acum += (sum(vec_ev)+obj1.valor)*0.10
                    
    if ddd.reten_11 =='SI':# calcular y ubicar a retencion del 11 %
        p.drawString(305,y-base_alt-(contadorr-1)*10-((contador + 4 + contt + ccont)*10), "23651502")
        p.drawString(365,y-base_alt-(contadorr-1)*10-((contador + 4 + contt + ccont)*10), "Retención 11%")
        p.drawString(500,y-base_alt-(contadorr-1)*10-((contador + 4 + contt + ccont)*10),"$ " + str('{:,}'.format(round(float((sum(vec_ev)+obj1.valor)*0.11)))))
        ccont += 1 
        acum += (sum(vec_ev)+obj1.valor)*0.11
        
    if ddd.reten_art_383 == 'SI':# calcular y ubicar la retencion del art 383
        p.drawString(305,y-base_alt-(contadorr-1)*10-((contador + 4 + contt + ccont)*10),"23651503")
        p.drawString(365,y-base_alt-(contadorr-1)*10-((contador + 4 + contt + ccont)*10),"Retención Art. 383")
        p.drawString(500,y-base_alt-(contadorr-1)*10-((contador + 4 + contt + ccont)*10), "$ " + str('{:,}'.format(re_valor_reten_cal)) )
        ccont += 1
        acum += re_valor_reten_cal
        
    if ddd.reten_arrindo == 'SI':# calcular y ubicar la retencion por concepto de ARRIENDO
        reten_provi = cuenta_reten.objects.get(contrato=ddd.id)
        # NOTA: arrindo es arriendo, es que me equivoque al escribirlo y lo deje asi
        ultimo_tipo_fac = bbb.last()

        p.drawString(305,y-base_alt-(contadorr-1)*10-((contador + 4 + contt + ccont)*10), str(reten_provi.num_cuenta))
        p.drawString(365,y-base_alt-(contadorr-1)*10-((contador + 4 + contt + ccont)*10),reten_provi.name_cuenta)
        if aaa.count() != 0:
            for i in aaa:
                if i.tipo_fact.id == ultimo_tipo_fac.id:
                    valor = i.valor
        else:
            valor = 0
                    
        p.drawString(500,y-base_alt-(contadorr-1)*10-((contador + 4 + contt + ccont)*10), "$ " + str('{:,}'.format((reten_provi.porc_retencion/100)*valor) ) )
        ccont += 1 
        acum += (reten_provi.porc_retencion/100)*valor


    p.line(500,y-base_alt-(contadorr-1)*10-((contador + 4 + contt + ccont)*10)+8, 550,y-base_alt-(contadorr-1)*10-((contador + 4 + contt + ccont)*10)+8)
    p.drawString(443,y-base_alt-(contadorr-1)*10-((contador + 4 + contt + ccont)*10), "Total Retención:")
    p.drawString(500,y-base_alt-(contadorr-1)*10-((contador + 4 + contt + ccont)*10), "$ " + str('{:,}'.format(round(float(acum)))))
        
    p.setFont('Helvetica-Bold', 6)
    p.drawString(446,y-base_alt-(contadorr-1)*10-((contador + 4 + contt + ccont+1)*10), "HONORARIO: ")
    p.drawString(500,y-base_alt-(contadorr-1)*10-((contador + 4 + contt + ccont+1)*10), "$ "+str('{:,}'.format(round(float(sum(resul)-sum(vec_mf)+obj1.valor-acum)))) )
    p.setFont('Helvetica', 6)
    
    #PIE DE PAGINA =================================================================================================
    p.line(40,70, 550,70)
    p.drawString(200,60, "CUENTA POR PAGAR GENERADA POR SICOS")
    p.drawString(245,50 , "Fundación Clínia del Río")
    p.drawString(191,40, "Dirección: Cra. 3 #128, Montería, Córdoba, Colombia")
    p.drawString(238,30, "Telefono: +57 311 7623443")
    
    
    # Captura usuario actual del Equipo
    current_user = request.user
    if User.is_active: 
        usuarioEquipo = getpass.getuser()  
        p.drawString(100,20, "Esta Factura fue Realizada en el Equipo " + '"'+str(usuarioEquipo)+'"'+", Por el Usuario " + '"' + str(current_user) + '"'+ ' el dia '+ time.strftime("%a-%d/%m/%y")+' a las '+ time.strftime(" %r"))

        
    #FIN PIE DE PAGINA =================================================================================================

    p.showPage()
    p.save()
    p=buffer.getvalue()
    buffer.close()
    response.write(p)
    return response

#reporte de los proveedores
def Reporte_Proveedor_Pdf(request, id_):
    response = HttpResponse(content_type='application/pdf')
    buffer=io.BytesIO()
    archivo_imagen = settings.STATICFILES_URL + '/img/clinicario_log.jpeg'
    p=canvas.Canvas(buffer, pagesize=A4)
    
    p.setFont('Helvetica', 7) 
    
    #ENCABEZADO ==============================================================================
    y1=750# ESTE ES EL PUNTO DE REFRENCIA DEL ENCABEZADO Y DE TODO EL DOUMENTO
    p.drawImage(archivo_imagen,40,y1-8, 95,65, preserveAspectRatio=True)
    p.drawString(240, y1+45, "FUNDACIÓN CLÍNICA DEL RÍO")
    p.drawString(260, y1+35, "NIT. 900 540 156-1")
    p.drawString(200, y1+20, "DISTRIBUCIÓN DE COSTOS Y GASTOS POR PAGAR POR ")
    p.drawString(235, y1+10, "PROVEEDORES CONTRATADOS")
    
    p.drawString(460, y1+45, "Código:"), p.drawString(505, y1+45, "xxxx")
    p.drawString(460, y1+35, "Version:"), p.drawString(505, y1+35, "xxxx")
    p.drawString(460, y1+25, "Emision:"), p.drawString(505, y1+25, "xxxx")
    p.drawString(460, y1+15, "Fecha:"), p.drawString(505, y1+15, time.strftime("%d/%m/%y"))
    #FIN ENCABEZADO ==============================================================================


    #DATOS
    y = y1-30 #punto de referencia para mover todos los objetos
    cpp_prov = cpp_proveedor.objects.get(id = id_)
    cc = centro_costo.objects.all()
    cuenta_auxx = cuenta_aux.objects.filter(proveedor = cpp_prov.proveedor.id)
    distri = distribucion.objects.filter(proveedor = cpp_prov.proveedor)
    cpp_prov_detal = cpp_proveedor_detalle.objects.filter(cpp_proveedor = id_)
    cont_cpp_prov_detal = cpp_prov_detal.count()
    categ = categoria.objects.all() 
    
    cpp_prov_calcular = defprov(cc,distri,cpp_prov_detal,cpp_prov,categ)
    
    (resultado, iva_matriz, distrib_mat, mat_detalle) = cpp_prov_calcular.calcular_cpp_proveedor()


    
    p.setFont('Helvetica', 8)
    
    #INICIO CUERPO DEL PDF===============================
    p.line(40,y+25, 550,y+25)
    p.drawString(40, y+16, "N° Factura: "), p.drawString(95, y+16, str(id_))
    p.drawString(40, y+8, "Proveedor:"), p.drawString(95, y+8, cpp_prov.proveedor.name_prov)
    p.drawString(40, y, "Cédula-NIT:"), p.drawString(95, y,cpp_prov.proveedor.nit_prov)
    p.line(40,y-2, 550,y-2)
    
    if cont_cpp_prov_detal != 0:#debe haber por lo menos un registro de cpp_proveedor_detalle
        
        #colocar los nombres de los prodcutos en la cabecera de la tabla
        divis = 350/(cont_cpp_prov_detal+1)#dividir la hoja en tamaños iguales
        
        p.drawString(260, y-12, "DISTRIBUCIONES")
        p.line(260,y-13, 328,y-13)

        if categ.first().id == distri.last().producto.categoria.id:#ESTO ES UN GAS
            
            
            p.line(40,y-25, 200,y-25)
            p.drawString( 40, y-33, "TABLA DISTRIBUCIÓN PORCENTUAL" )
            p.line(40,y-35, 550,y-35)
            
            y = y-20
            
            #colocar los nombres de los centros de costo en las filas de la tabla
            p.drawString( 40, y-23, "Centro de Costo" )
            p.line(40,y-24, 115,y-24)
            
            p.setFont('Helvetica', 6)
            varx = 0
            for i in cpp_prov_detal:
                p.drawString((varx*divis)+200, y-23, i.producto.nombre)
                p.line((varx*divis)+200,y-24, (varx*divis)+200 +divis-20,y-24)
                varx += 1
            
            p.setFont('Helvetica', 8)
            #escribir los nombres de los centros de costos
            vary = 0
            for j in cc:
                p.drawString( 40, y-33-(vary*8),j.name_ccos )
                vary += 1
                
            for i in range(cc.count()):
                for j in range(cont_cpp_prov_detal):
                    if distrib_mat[i,j] != 0:
                        p.drawString( (j*divis)+200 , y-33-(i*8) , str(round(float(distrib_mat[i,j]))) + '%' )
                    else:
                        p.drawString( (j*divis)+200 , y-33-(i*8) , '-' )
                        
            p.line(40,y-33-(i*8)-2, 550,y-33-(i*8)-2)
            
            
            #OTRA tabla
            y = y-33-(i*8)#actualizar el valor de "y"
            
            p.line(40,y-25, 200,y-25)
            p.drawString( 40, y-33, "TABLA DETALLE" )
            p.line(40,y-35, 550,y-35)
            
            #colocar los nombres de los centros de costo en las filas de la tabla
            p.drawString( 40, y-43, "Producto" ), p.drawString( 200, y-43, "Cantidad" ),  p.drawString( 245, y-43, "Cant. Cilindro" ), p.drawString( 310, y-43, "Valor Unit." ), p.drawString( 370, y-43, "Subtotal" ), p.drawString( 440, y-43, "Iva" ), p.drawString( 500, y-43, "¿Maneja Iva?" )
            p.line(40,y-44, 115,y-44)   ,         p.line(200,y-44, 235,y-44),             p.line(245,y-44, 295,y-44),                     p.line(310,y-44, 360,y-44),            p.line(370,y-44, 420,y-44),             p.line(440,y-44, 480,y-44),       p.line(500,y-44, 550,y-44)
            
            ccont = 0
            for i in cpp_prov_detal:
                buscar_iva = distri.filter(producto = i.producto.id).last()
                
                p.drawString( 40, y-52-(ccont*8), i.producto.nombre)#escribir el nombre del producto
                p.drawString( 200, y-52-(ccont*8), str(int(i.cant_produc)))#escribir cantidad del producto
                
                if i.cant_flete*i.valor_flete != 0:#escribir el valor del impuesto de transporte (flete)
                    p.drawString( 245, y-52-(ccont*8),'$ ' + str(i.cant_flete*i.valor_flete))
                else:
                    p.drawString( 245, y-52-(ccont*8),'-')
                    
                p.drawString( 310, y-52-(ccont*8), '$ ' + str('{:,}'.format(round(float(i.producto.precio)))))#escribir el valor del producto
                p.drawString( 370, y-52-(ccont*8), '$ ' + str('{:,}'.format(round(float(i.producto.precio*i.cant_produc)))))#escribir el valor total del producto
                p.drawString( 440, y-52-(ccont*8), '$ ' + str('{:,}'.format(round(float(sum(iva_matriz[0:cc.count(), ccont]))))))#escribir el valor del iba dle producto correspondiente               
                p.drawString( 520, y-52-(ccont*8), buscar_iva.meneja_iva)#escribir si el producto maneja iva o no       

                ccont += 1
            
            p.line(40,y-52-(ccont*8)+6, 550,y-52-(ccont*8)+6)
            
            
            p.line(40,y-52-(ccont*8)-20, 200,y-52-(ccont*8)-20)
            p.drawString( 40, y-52-(ccont*8)-28, "TABLA DISTRIBUCIÓN CONTABLE" )
            p.line(40,y-52-(ccont*8)-30, 550,y-52-(ccont*8)-30)
            
            p.drawString( 40,y-52-(ccont*8)-38, "Centro Costo" ), p.drawString( 200,y-52-(ccont*8)-38, "Cuenta" ),       p.drawString( 287,y-52-(ccont*8)-38, "Costo" ),         p.drawString( 375,y-52-(ccont*8)-38, "Iva" ),          p.drawString( 460,y-52-(ccont*8)-38, "Cuenta Iva" ),       
            p.line(40,y-52-(ccont*8)-40, 115,y-52-(ccont*8)-40),  p.line(200,y-52-(ccont*8)-40, 255,y-52-(ccont*8)-40),  p.line(287,y-52-(ccont*8)-40, 342,y-52-(ccont*8)-40),   p.line(375,y-52-(ccont*8)-40, 430,y-52-(ccont*8)-40),  p.line(460,y-52-(ccont*8)-40, 515,y-52-(ccont*8)-40)
            
            contn = 0
            for k in cc:
                p.drawString( 40,y-52-(ccont*8)-48-(contn*8), k.name_ccos)#ESCRIBIR CENTRO DE COSTO 
                
                if distrib_mat[contn,cont_cpp_prov_detal] == 0:#escribir la cuenta contable del costo
                    p.drawString( 200,y-52-(ccont*8)-48-(contn*8), '-')
                else:
                    p.drawString( 200,y-52-(ccont*8)-48-(contn*8), str(int(distrib_mat[contn,cont_cpp_prov_detal])))
                
                if resultado[contn,cont_cpp_prov_detal] != 0:#escribir el costo
                    p.drawString( 287,y-52-(ccont*8)-48-(contn*8),'$ ' + str('{:,}'.format(round(float(resultado[contn,cont_cpp_prov_detal])))))
                    p.drawString( 375,y-52-(ccont*8)-48-(contn*8),'$ ' + str('{:,}'.format(round(float(iva_matriz[contn,cont_cpp_prov_detal])))))
                    p.drawString( 460,y-52-(ccont*8)-48-(contn*8),str(round(float(iva_matriz[contn,cont_cpp_prov_detal+1]))))                                        
                else:
                    p.drawString( 287,y-52-(ccont*8)-48-(contn*8),'-') 
                    p.drawString( 375,y-52-(ccont*8)-48-(contn*8),'-')                    
                    p.drawString( 460,y-52-(ccont*8)-48-(contn*8),'-') 

                contn += 1
     
            p.line(40,y-52-(ccont*8)-48-(contn*8)+6, 550,y-52-(ccont*8)-48-(contn*8)+6)
     
     
            #actualizar el valor de "y"
            y = y-52-(ccont*8)-48-(contn*8)-15
            
            p.line(40,y, 200,y)
            p.drawString( 40, y-8, "TABLA DISTRIBUCIÓN DÉBITO-CRÉDITO" )
            p.line(40,y-10, 550,y-10)
            
            p.drawString( 40,y-18, "Centro Costo" ),        p.drawString( 287,y-18, "Débito" ),         p.drawString( 375,y-18, "Crédito" )           
            p.line(40,y-20, 115,y-20),                                  p.line(287,y-20, 342,y-20),                    p.line(375,y-20, 430,y-20)           
            # p.drawString( 200,y-18, "Cuenta" ),
            # p.line(200,y-20, 255,y-20),



            vary = 0
            for j in cc:
                p.drawString( 40,y-28-(8*vary),j.name_ccos )
                
                if distrib_mat[vary,cont_cpp_prov_detal] == 0:#escribir la cuenta contable del costo
                    # p.drawString( 200,y-28-(8*vary), '-')
                    p.drawString( 287,y-28-(8*vary),'-')    
                    p.drawString( 375,y-28-(8*vary),'-')    
                                
                else:
                    # p.drawString( 200,y-28-(8*vary), str(int(distrib_mat[vary,cont_cpp_prov_detal])))
                    p.drawString( 287,y-28-(8*vary),'$ ' + str('{:,}'.format(round(float(resultado[vary,cont_cpp_prov_detal]+iva_matriz[vary,cont_cpp_prov_detal])))))
                    p.drawString( 375,y-28-(8*vary),'-')
                    
                vary += 1
                
                
            #se calcula la cantidad de plata a retener
            total = sum(resultado[0:cc.count(),cont_cpp_prov_detal])+ sum(iva_matriz[0:cc.count(),cont_cpp_prov_detal])
            
            conta = 0
            for m in cuenta_auxx:
                if cpp_prov.aire_medicinal == 'SI':
                    if m.cuenta[0:2] == '23':
                        p.drawString( 40,y-28-(8*vary), m.name_cuenta)
                        # p.drawString( 200,y-28-(8*vary), m.cuenta)
                        p.drawString( 287,y-28-(8*vary),'-')
                        p.drawString( 375,y-28-(8*vary),'$ ' + str('{:,}'.format(total)))
                        
                else:
                    if m.cuenta[0:2] == '22':
                        p.drawString( 40,y-28-(8*vary), m.name_cuenta)
                        # p.drawString( 200,y-28-(8*vary), m.cuenta)
                        p.drawString( 287,y-28-(8*vary),'-')
                        p.drawString( 375,y-28-(8*vary),'$ ' + str('{:,}'.format(total)))
            
            p.drawString(40,y-28-(8*vary)-10,'TOTAL')
            # p.drawString(200,y-28-(8*vary)-10,'-')
            
            p.line(287,y-28-(8*vary)-2,342, y-28-(8*vary)-2),                              p.line(375,y-28-(8*vary)-2,430, y-28-(8*vary)-2)
            p.drawString( 287,y-28-(8*vary)-10,'$ ' + str('{:,}'.format(total))),          p.drawString( 375,y-28-(8*vary)-10,'$ ' + str( '{:,}'.format(total)))
            
            
                          
        else:#ESTO ES SANGRE
            
            p.drawString( 40, y-25, "VALOR DEL FLETE: " + '$ ' +str('{:,}'.format(cpp_prov_detal.last().valor_flete)))
            
            y = y-25
            p.line(40,y-6, 200,y-6)
            p.drawString( 40, y-14, "TABLA DISTRIBUCIÓN CONTABLE" )
            p.line(40,y-15, 550,y-15)
            
            #colocar los nombres de los centros de costo en las filas de la tabla
            p.drawString( 40, y-23, "Centro de Costo" )
            p.line(40,y-24, 115,y-24)
            
            #PRIMERA TABLA======================================================================
            #escribir los nombres de los centros de costos
            vary = 0
            for j in cc:
                p.drawString( 40, y-33-(vary*8),j.name_ccos )
                vary += 1
                
            #colocar los nombres de los centros de costo en las filas de la tabla
            p.drawString( 200, y-23, "Distribución Final" ) ,  p.drawString( 300, y-23, "Cuenta" ),   p.drawString( 400, y-23, "Débito" ), p.drawString( 485, y-23, "Crédito" )
            p.line(200,y-24, 262,y-24) ,                       p.line(300,y-24, 350,y-24),           p.line(400,y-24, 450,y-24),            p.line(485,y-24, 550,y-24)
            
            #recorrer los centrros de costos
            for i in range(cc.count()):
                if distrib_mat[i,0] != 0:
                    p.drawString( 200, y-33-(i*8) , str(round(float(distrib_mat[i,0])))+ '%' ) 
                    p.drawString( 300, y-33-(i*8) , str(round(float(iva_matriz[i,1]))))
                    p.drawString( 400, y-33-(i*8) , '$ ' + str('{:,}'.format(round(float(resultado[i,0]))) ))
                    p.drawString( 485, y-33-(i*8) , '-' )
                    
                else:
                    p.drawString( 200, y-33-(i*8), '-' )
                    p.drawString( 300, y-33-(i*8) , '-')
                    p.drawString( 400, y-33-(i*8), '-' )
                    p.drawString( 485, y-33-(i*8) , '-' )
                    
            
            #recorrer las cuentas auxiliares
            #se calcula la cantidad de plata a retener
            total = sum(resultado[0:cc.count(),0])
            can_reten = total*cpp_prov.reten/100
            
            recor = 0
            for r in cuenta_auxx:
                if r.cuenta[0:4] == '2365':
                    p.drawString( 40, y-33-(i*8)-(recor*8) -8, r.name_cuenta) 
                    p.drawString( 200, y-33-(i*8)-(recor*8) -8, '-')
                    p.drawString( 400, y-33-(i*8)-(recor*8) -8, '-')
                    p.drawString( 485, y-33-(i*8)-(recor*8)-8,'$ ' + str('{:,}'.format(round(can_reten) )))
                    p.drawString( 300, y-33-(i*8)-(recor*8)-8, r.cuenta )
    
                else:
                    p.drawString( 40, y-33-(i*8)-(recor*8)-8, r.name_cuenta) 
                    p.drawString( 200, y-33-(i*8)-(recor*8) -8, '-') 
                    p.drawString( 400, y-33-(i*8)-(recor*8) -8, '-')
                    p.drawString( 485, y-33-(i*8)-(recor*8)-8,'$ ' + str('{:,}'.format(round(total-can_reten) )))
                    p.drawString( 300, y-33-(i*8)-(recor*8)-8, r.cuenta )

                    
                recor += 1
            
            #aqui se coloca el valor total de la cpp  debito y credito    
            p.drawString( 40, y-33-(i*8)-(recor*8)-8,'TOTAL' )
            
            p.line(485,y-33-(i*8)-(recor*8)-1,550,y-33-(i*8)-(recor*8)-1)
            p.drawString( 485, y-33-(i*8)-(recor*8)-8,'$ ' + str('{:,}'.format(round(float(np.sum(resultado)))) ))
            
            p.line(400,y-33-(i*8)-(recor*8)-1,450,y-33-(i*8)-(recor*8)-1)
            p.drawString( 400, y-33-(i*8)-(recor*8)-8,'$ ' + str('{:,}'.format(round(float(np.sum(resultado)))) ))
            
            p.line(40,y-33-(i*8)-(recor*8)-10,550,y-33-(i*8)-(recor*8)-10)

            #==================================================================FIN PRIMERA TABLA


            #segunda     TABLA======================================================================
            var_filas = y-33-(i*8)-60
        
            p.line(40,var_filas+6, 200,var_filas+6)
            p.drawString( 40, var_filas-2, "TABLA DISTRIBUCIÓN GLOBAL" )
            p.line(40,var_filas-3, 550,var_filas-3)
            
            #colocar los nombres de los centros de costo en las filas de la tabla
            p.drawString( 40, var_filas-11, "Centro de Costo" )
            p.line(40,var_filas-12, 115,var_filas-12)
            
            varx = 0
            for i in cpp_prov_detal:
                p.drawString((varx*divis)+200, var_filas-11, i.producto.nombre)
                p.line((varx*divis)+200,var_filas-12, (varx*divis)+200 +divis-20,var_filas-12)
                varx += 1
                
            p.drawString((varx*divis)+200, var_filas-11, 'TOTAL')
            p.line((varx*divis)+200,var_filas-12, (varx*divis)+200 +divis-20,var_filas-12)

            
            vary = 0
            for j in cc:
                p.drawString( 40,var_filas-20-(vary*8),j.name_ccos )

                varx = 0
                for k in range(cont_cpp_prov_detal+1):
                    if mat_detalle[vary,varx] != 0:
                        p.drawString( (varx*divis) + 200, var_filas-20-(vary*8),'$ '+ str('{:,}'.format(round(mat_detalle[vary,varx]))))
                    else:
                        p.drawString( (varx*divis) + 200, var_filas-20-(vary*8),'-')
                       
                    
                    varx += 1

                vary += 1   
                
            p.line(40,var_filas-20-(vary*8),550,var_filas-20-(vary*8))
    
            #==================================================================FIN segunda TABLA
            
            
            
            vary2 = var_filas-20-(vary*8)-35
            p.line(40,vary2+8, 200,vary2+8)
            p.drawString( 40, vary2, "TABLA DISTRIBUCIÓN UNITARIA" )
            p.line(40,vary2-2, 550,vary2-2)

            p.drawString( 40, vary2-10, "Producto" ), p.drawString( 142, vary2-10, "Centro Costo" ), p.drawString( 224, vary2-10, "Valor Unitario" ),p.drawString( 306, vary2-10, "Cantidad" ), p.drawString( 368, vary2-10, "Flete" ), p.drawString( 450, vary2-10, "Total" )
            p.line(40,vary2-11, 100,vary2-11),        p.line(142,vary2-11, 202,vary2-11),            p.line(224,vary2-11, 284,vary2-11),             p.line(306,vary2-11, 350,vary2-11),        p.line(368,vary2-11, 428,vary2-11),     p.line(450,vary2-11, 550,vary2-11)
            
            esp_y = 0
            for i in cpp_prov_detal:
                p.drawString( 40, vary2-19-(8*esp_y), i.producto.nombre)
                p.drawString( 142, vary2-19-(8*esp_y), str(i.centro_costo.name_ccos))
                p.drawString( 224, vary2-19-(8*esp_y), '$ ' + str('{:,}'.format(i.valor_produc)))
                p.drawString( 306, vary2-19-(8*esp_y), str('{:,}'.format(int(i.cant_produc))))
                p.drawString( 368, vary2-19-(8*esp_y), '$ ' + str('{:,}'.format(round(float(i.valor_flete/cont_cpp_prov_detal),2))))
                p.drawString( 450, vary2-19-(8*esp_y), '$ ' + str('{:,}'.format(round(float((i.cant_produc*i.valor_produc)+i.valor_flete/cont_cpp_prov_detal),2))))

                esp_y += 1
            
    #=============================FIN DEL CUERPO DEL PDF
        
        
    #PIE DE PAGINA =================================================================================================
    p.line(40,70, 550,70)
    p.drawString(200,60, "CUENTA POR PAGAR GENERADA POR SICOS")
    p.drawString(245,50 , "Fundación Clínia del Río")
    p.drawString(191,40, "Dirección: Cra. 3 #128, Montería, Córdoba, Colombia")
    p.drawString(238,30, "Telefono: +57 311 7623443")
    
    # Captura usuario actual del Equipo
    current_user = request.user
    if User.is_active: 
        usuarioEquipo = getpass.getuser()  
        p.drawString(70,20, "Esta Factura fue Realizada en el Servidor " + '"'+str(usuarioEquipo)+'"'+", Por el Usuario " + '"' + str(current_user) + '"'+ ' el dia '+ time.strftime("%a-%d/%m/%y")+' a las '+ time.strftime(" %r"))
    
    #==================================FIN PIE DE PAGINA

    p.showPage()
    p.save()
    # buffer.seek(0)
    p=buffer.getvalue()
    buffer.close()
    response.write(p)
    return response

#reporte de los Servicios
def Reporte_servicios_Pdf(request, id_):
    
    response = HttpResponse(content_type='application/pdf')
    buffer=io.BytesIO()
    archivo_imagen = settings.STATICFILES_URL + '/img/clinicario_log.jpeg'
    p=canvas.Canvas(buffer, pagesize=A4)
    
    p.setFont('Helvetica', 7) 
    
    #ENCABEZADO ==============================================================================
    y1=750# ESTE ES EL PUNTO DE REFRENCIA DEL ENCABEZADO Y DE TODO EL DOUMENTO
    p.drawImage(archivo_imagen,40,y1-8, 95,65, preserveAspectRatio=True)
    p.drawString(240, y1+45, "FUNDACIÓN CLÍNICA DEL RÍO")
    p.drawString(260, y1+35, "NIT. 900 540 156-1")
    p.drawString(200, y1+20, "DISTRIBUCIÓN DE COSTOS Y GASTOS POR PAGAR POR ")
    p.drawString(235, y1+10, "PROVEEDORES CONTRATADOS")
    
    p.drawString(460, y1+45, "Código:"), p.drawString(505, y1+45, "xxxx")
    p.drawString(460, y1+35, "Version:"), p.drawString(505, y1+35, "xxxx")
    p.drawString(460, y1+25, "Emision:"), p.drawString(505, y1+25, "xxxx")
    p.drawString(460, y1+15, "Fecha:"), p.drawString(505, y1+15, time.strftime("%d/%m/%y"))
    #FIN ENCABEZADO ==============================================================================
    
    #ESTE ES EL CUERPO DEL PDF=====================================================================
    y = y1-40 #punto de referencia para mover todos los objetos
    
    #DATOS
    cpp_serv_public_tem = cpp_serv_public.objects.get(id = id_)
    cpp_ser_pub = cpp_serv_public.objects.get(id = id_)
    distribu = distri_serv_public.objects.filter(serv_public = cpp_ser_pub.serv_public.id)
    calcular_distri = funcion_serv(cpp_ser_pub,distribu)
    vector_distri = calcular_distri.distri_serv()
    
    
    p.setFont('Helvetica', 8)
    
    #INICIO CUERPO DEL PDF===============================
    p.line(40,y+25, 550,y+25)
    p.drawString(40, y+16, "N° Factura: "), p.drawString(115, y+16, str(id_))
    p.drawString(40, y+8, "Tipo de Servicio:"), p.drawString(115, y+8,cpp_serv_public_tem.serv_public.tipo_serv.nombre_tipo )
    p.drawString(40, y, "Cédula-NIT:"), p.drawString(115, y, cpp_serv_public_tem.serv_public.nit)
    p.drawString(40, y-8, "Nombre del Tercero:"), p.drawString(115, y-8, cpp_serv_public_tem.serv_public.nombre_tercero)
    p.drawString(40, y-16, "Nombre del Servicio:"), p.drawString(115, y-16, cpp_serv_public_tem.serv_public.nombre_serv)
    p.line(40,y-18, 550,y-18)
    
    p.drawString(265,y-30, "DISTRIBUCIÓN")
    p.line(265,y-32, 323,y-32)
    p.line(265,y-34, 323,y-34)
     
    y = y-34 # punto de referencia
    
    #primera tabla, colocar los porcentajes de las distribuciones dentro de cada centro de costo.
    p.line(40,y-11, 200,y-11)
    p.drawString( 40, y-20, "TABLA DISTRIBUCIÓN PORCENTUAL" )
    p.line(40,y-22, 550,y-22)
    
    #Campos de la primera tabla
    p.drawString( 40, y-32, "Centro de Costo" ), p.drawString( 200, y-32, "Porcentaje" ),
    p.line(40,y-34, 115,y-34)   ,         p.line(200,y-34, 240,y-34), 
    
    inter = 0
    for i in distribu:
        p.drawString( 40, y-42-(inter*8), i.centro_costo.name_ccos)   
        p.drawString( 200, y-42-(inter*8), str(i.distri) + "%")
        
        inter += 1
        
        
    detalle = cpp_servi_detalle.objects.filter(cpp_serv_public = id_) 
     
    if detalle.count() == 0:#no hay registros almacenados en el cpp_servi_detalle    
        y = y-42-(inter*8)#Punto de referencia
        
        #esta es la segunda tabla
        p.line(40,y-10, 200,y-10)
        p.drawString( 40, y-20, "TABLA DISTRIBUCIÓN CONTABLE" )
        p.line(40,y-22, 550,y-22)
        
        #campos de la segunda tabla
        p.drawString( 40, y-32, "Centro de Costo" ),  p.drawString( 200, y-32, "Cuenta" ),   p.drawString( 270, y-32, "Costo" ),   p.drawString( 340, y-32, "Cuenta Iva" ),  p.drawString( 410, y-32, "Valor Iva" ), p.drawString( 480, y-32, "Crédito" )
        p.line(40,y-34, 115,y-34)   ,                 p.line(200,y-34, 240,y-34),            p.line(270,y-34, 310,y-34),           p.line(340,y-34, 380,y-34),               p.line(410,y-34, 450,y-34),             p.line(480,y-34, 530,y-34)

        
        inter2 = 0
        
        #llenar tabla
        for i in distribu:
            p.drawString( 40, y-42-(inter2*8), i.centro_costo.name_ccos)
            p.drawString( 200, y-42-(inter2*8), i.num_cuenta_especific)
            p.drawString( 270, y-42-(inter2*8), "$ " + str('{:,}'.format(vector_distri[inter2,0])))
            p.drawString( 340, y-42-(inter2*8), i.num_cuenta_iva)
            p.drawString( 410, y-42-(inter2*8), "$ " + str('{:,}'.format(vector_distri[inter2,1])))
            p.drawString( 480, y-42-(inter2*8), "-")

            inter2 += 1

        #recorrer las cuentas contra
        cuenta_aux_servicios = cuenta_aux_serv.objects.filter(serv_public = cpp_serv_public_tem.serv_public)

        
        inter3 = 0
        
        #llenar tabla con las cuentas contra
        for t in cuenta_aux_servicios:
            p.drawString( 40, y-42-(inter2*8)-(inter3*8), t.name_cuenta)
            p.drawString( 200, y-42-(inter2*8)-(inter3*8), t.num_cuenta)
            p.drawString( 270, y-42-(inter2*8)-(inter3*8), "-")
            p.drawString( 340, y-42-(inter2*8)-(inter3*8), "-")
            p.drawString( 410, y-42-(inter2*8)-(inter3*8), "-")
            
            if t.num_cuenta[0:4] == "2365":#esto es una retencion
                p.drawString( 480, y-42-(inter2*8)-(inter3*8), "$ "+str('{:,}'.format(cpp_ser_pub.costo*(cpp_ser_pub.reten/100))))
            else:#esto es el total de la CPP
                p.drawString( 480, y-42-(inter2*8)-(inter3*8), "$ "+str('{:,}'.format(np.sum(vector_distri)-(cpp_ser_pub.costo*(cpp_ser_pub.reten/100)))))

            inter3 += 1
    else:#si hay registros dentro del detalle
        
        y = y-42-(inter*8)#Punto de referencia
        
        #esta es la segunda tabla
        p.line(40,y-10, 200,y-10)
        p.drawString( 40, y-20, "TABLA DISTRIBUCIÓN CONTABLE" )
        p.line(40,y-22, 550,y-22)
        
        #campos de la segunda tabla
        p.drawString( 40, y-32, "Centro de Costo" ),  p.drawString( 200, y-32, "Cuenta" ),   p.drawString( 270, y-32, "Costo" ),   p.drawString( 340, y-32, "Cuenta Iva" ),  p.drawString( 410, y-32, "Valor Iva" ), p.drawString( 480, y-32, "Crédito" )
        p.line(40,y-34, 115,y-34)   ,                 p.line(200,y-34, 240,y-34),            p.line(270,y-34, 310,y-34),           p.line(340,y-34, 380,y-34),               p.line(410,y-34, 450,y-34),             p.line(480,y-34, 530,y-34)

        inter2 = 0
        
        #llenar tabla
        for i in detalle:
            
            p.drawString( 200, y-42-(inter2*8), i.num_cuenta_aux)
            
            if i.id_centro_costo == 0:
                p.drawString( 40, y-42-(inter2*8), i.name_cuenta_aux)
                p.drawString( 270, y-42-(inter2*8), "-")
                p.drawString( 340, y-42-(inter2*8), "-")
                p.drawString( 410, y-42-(inter2*8), "-")
                p.drawString( 480, y-42-(inter2*8), "$ " + str('{:,}'.format(i.valor_contra)))
            else:
                p.drawString( 40, y-42-(inter2*8), i.name_ccos)
                p.drawString( 270, y-42-(inter2*8), "$ " + str('{:,}'.format(i.costo)))
                p.drawString( 340, y-42-(inter2*8), str(i.num_cuenta_iva))
                p.drawString( 410, y-42-(inter2*8), "$ " + str('{:,}'.format(i.valor_iva)))
                p.drawString( 480, y-42-(inter2*8), "-")
  
            inter2 += 1

        
        
        
            


        
    
    
    
    #ESTE ES EL FIN DEL CUERPO DEL PDF=====================================================================

    
        
        
    #PIE DE PAGINA =================================================================================================
    p.line(40,70, 550,70)
    p.drawString(200,60, "CUENTA POR PAGAR GENERADA POR SICOS")
    p.drawString(245,50 , "Fundación Clínia del Río")
    p.drawString(191,40, "Dirección: Cra. 3 #128, Montería, Córdoba, Colombia")
    p.drawString(238,30, "Telefono: +57 311 7623443")
    
    # Captura usuario actual del Equipo
    current_user = request.user
    if User.is_active: 
        usuarioEquipo = getpass.getuser()  
        p.drawString(70,20, "Esta Factura fue Realizada en el Servidor " + '"'+str(usuarioEquipo)+'"'+", Por el Usuario " + '"' + str(current_user) + '"'+ ' el dia '+ time.strftime("%a-%d/%m/%y")+' a las '+ time.strftime(" %r"))
    
    #==================================FIN PIE DE PAGINA

    p.showPage()
    p.save()
    # buffer.seek(0)
    p=buffer.getvalue()
    buffer.close()
    response.write(p)
    return response


#==================================================================================================
#|||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||
#==================================================================================================
#==========******VISTAS DEL HOME *****=============================================================
#==================================================================================================
#|||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||
#==================================================================================================


def index(request,*args, **kwargs):
    return render(request, 'home.html', {})

def base(request,*args, **kwargs):
    return render(request, 'base.html', {})

def base_cirugia(request,*args, **kwargs):
    return render(request, 'base_cirugia.html', {})

def cambiar_pass_form(request,*args, **kwargs):
    return render(request, 'usuario/cambiar_pass_form.html', {})

def cambiar_pass_form_error(request,*args, **kwargs):
    return render(request, 'usuario/cambiar_pass_form_error.html', {})

def cambiar_pass(request,*args, **kwargs):
    usuario = request.GET.get("usuario")
    passw = request.GET.get("passw")
    passw_2 = request.GET.get("passw_2")
    
    if passw == passw_2:
        user = User.objects.get(username=usuario)
        user.set_password(passw)
        user.save()
        return redirect('login')
    else:
        return redirect('cambiar_pass_form_error')
    


#==================================================================================================
#|||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||
#==================================================================================================
#==========******VISTAS DE ESPECIALISTAS******=============================================================
#==================================================================================================
#|||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||
#==================================================================================================

#ESTA FUNCION ES PARA CREAR UNA LISTA DE LOS ESPECIALISTAS

def cuenta_reten_list(request):
    cuenta_reten1 = cuenta_reten.objects.all()
    contexto = {'cuenta_retens': cuenta_reten1}
    return render(request, 'especialistas/cuenta_reten_list.html',contexto)

class cuenta_retenEdit(UpdateView):
    model = cuenta_reten
    form_class = cuenta_retenform
    template_name = 'especialistas/cuenta_reten_form.html'
    success_url = reverse_lazy('cuenta_reten_list')

class cuenta_retenCrear(CreateView):
    model = cuenta_reten
    form_class = cuenta_retenform
    template_name = 'especialistas/cuenta_reten_form.html'
    success_url = reverse_lazy('cuenta_reten_list')


def especialistas_list(request):
    
    filtronombre=request.GET.get("filtronombre")#extrae de especialistas_list.html el valor de la entrada de texto llamada "filtroname" que corresponde al nombre del especialista
    def __str__(selt):
        return '%s ' % (self.filtronombre)


    filtroid=request.GET.get("filtroid")#extrae de especialistas_list.html el valor de la entrada de texto llamada "filtroid" que corresponde al nombre del especialista
    def __str__(selt):
        return '%s ' % (self.filtroid)

    
    if filtronombre:
        especialista1 = especialista.objects.filter(name_esp__icontains = filtronombre)#aqui se filtran los objetos que cumplen el filtr de nombre
       
        if filtroid:#aqui se filtra nuevamente pero por "id" a los especialistas que cumplieron con el filtro "nombre"
            especialista1 = especialista.objects.filter(id_esp__icontains =filtroid)       
        
        contexto = {'especialistas':especialista1}
        return render(request,'especialistas_list.html', contexto)
    else:#este else se hace para que el usuario pueda filtrar por cualquiera de los dos filtros ademas de los dos al mismo tiempo

        if filtroid:
            especialista1 = especialista.objects.filter(id_esp__icontains = filtroid)#primero se filtra por "id"
        
            if filtronombre:#despues se filtra por "nombre"
                especialista1 = especialista.objects.filter(name_esp__icontains =filtronombre)       
        
            contexto = {'especialistas':especialista1}
            return render(request,'especialistas_list.html', contexto)
        else:#si los filtros estan en blanco se muestran todos los registros
            especialista1 = especialista.objects.all()
            contexto = {'especialistas':especialista1}
            return render(request,'especialistas_list.html', contexto)

    
#ESTA CLASE ES PARA EDITA A LOS ESPECIALISTAS REGISTRADOS    
class especialistaEdit(UpdateView):
    model = especialista
    form_class = especialistaform
    template_name = 'especialista_form.html'
    success_url = reverse_lazy('especialistas_list')


#ESTA CLASE ES PARA CREAR A LOS ESPECIALISTAS
def especialistaCrear(request):
    if request.method == 'POST' :
        form = especialistaform(request.POST)
        if form.is_valid(): 
            form.save()
        return redirect('especialista_view')
    else:
        form = especialistaform()
    
    return render(request, 'especialista_form.html', {'form':form} )
#[[[[[[[[====FINALIZA ESPECIALISTAS ===[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]
#******************************************************************************************************************************************************************************


#ESTA CLASE ES PARA LOS TERMINOS DEL CONTRATO
def contrato_list(request):
    contrato1=contrato.objects.all()
    contexto = {'contratos':contrato1}
    return render(request, 'contrato/contrato_list.html', contexto)

class contratoCrear(CreateView):
    model = contrato
    form_class = contratoform
    template_name = 'contrato/contrato_form.html'
    success_url = reverse_lazy('contrato_list')
    
class contratoEditar(UpdateView):
    model = contrato
    form_class = contratoform
    template_name = 'contrato/contrato_form.html'
    success_url = reverse_lazy('contrato_list')


#ESTA CLASE ES PARA el UVT
def uvt_list(request):
    uvt1=uvt.objects.all()
    contexto={'uvts':uvt1}
    return render(request,'contrato/uvt_list.html',contexto)

class uvtCrear(CreateView):
    model = uvt
    form_class = uvtform
    template_name= 'contrato/uvt_form.html'
    success_url = reverse_lazy('uvt_list')
    
class uvtEditar(UpdateView):
    model = uvt
    form_class = uvtform
    template_name= 'contrato/uvt_form.html'
    success_url = reverse_lazy('uvt_list')
    
    
#ESTA CLASE ES PARA RETEN_383
def reten_383_list(request):
    reten_3831=reten_383.objects.all()
    contexto = {'reten_383s':reten_3831} 
    return render(request,'contrato/reten_383_list.html',contexto)

class reten_383Crear(CreateView):
    model = reten_383
    form_class= reten_383form
    template_name='contrato/reten_383_form.html'
    success_url = reverse_lazy('reten_383_list')

class reten_383Editar(UpdateView):
    model = reten_383
    form_class= reten_383form
    template_name='contrato/reten_383_form.html'
    success_url = reverse_lazy('reten_383_list')


class centro_actividadCrear(CreateView):
    model = centro_actividad
    form_class = centro_actividadform
    template_name = 'especialistas/centro_actividad_form.html'
    success_url = reverse_lazy('centro_actividad_list')
    
    
class centro_actividadEdit(UpdateView):
    model=centro_actividad
    form_class= centro_actividadform
    template_name = 'especialistas/centro_actividad_form.html'
    success_url =reverse_lazy('centro_actividad_list')


def centro_actividad_list(request):
    centro_actividad1= centro_actividad.objects.all()
    contexto = {'centro_actividads': centro_actividad1}
    return render(request,'especialistas/centro_actividad_list.html', contexto)

#[[[[[[[[[====INICIA TIPOS DE FACTURACION===========[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]
#**************************************************************************************************************************************************************************
#ESTA CLASE ES PARA REGISTRAR LOS TIPOS DE FACTURACION (MIXTO, EVENTO, turno)
class tipo_faccrear(CreateView):
    model=tipo_fact  
    form_class = tipo_factform 
    template_name = 'tipo_factform.html'
    success_url = reverse_lazy('tipo_fact_list') 


class tipo_factEditar(UpdateView):
    model=tipo_fact
    form_class=tipo_factform
    template_name='tipo_factform.html'
    success_url=reverse_lazy('tipo_fact_list')


#ESTA FUNCION LISTA LOS REISTROS TIPOS DE FACTURA
def tipo_fact_list(request):
    tipo_fac1= tipo_fact.objects.all()
    contexto = {'tipo_facts': tipo_fac1}
    return render(request,'tipo_fact_list.html', contexto)
#[[[[[[[[[====FINALIZA TIPOS DE FACTURACION============[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]
#**************************************************************************************************************************************************************************





#[[[[[[[[[====INICIA FAC_ESPECIALISTAS_DETALLE==========[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]
#**************************************************************************************************************************************************************************
#ESTA FUNCION ES PARA REGISTRAR La fcaturacion detallada
def fac_especialista_detallecrear(request, id_):
    
    contar_detalle_fact=fac_especialista_detalle.objects.count()
    filtro=fac_especialista_detalle.objects.filter(fac_especialista = id_)
    filtro_cont = filtro.count()
    
    filtro1=fac_especialista_detalle.objects.filter(fac_especialista=id_)
    fac_especialista_detalle1=fac_especialista_detalle.objects.filter(fac_especialista=id_)
    
    tipos_facturacion= tipo_fact.objects.all()
    ccc = centro_actividad.objects.all()
    
    tempo =fac_especialista.objects.get(id=id_)
    ddd = contrato.objects.get(especialista = tempo.especialista.id)
    
    espe_cpp = factura_pdf(filtro1,tipos_facturacion, ccc, ddd )
    (resul, total_sum,a,b,c,d,e,f,g)= espe_cpp.opera_especialistas()
    

    if request.method == 'POST' :
        form = fac_especialista_detalleform(request.POST)
        if form.is_valid(): 

            form.save()
        return redirect('/crearfacdetal/'+str(id_)+'/')
    else:
        if filtro_cont != 0:
            a = filtro.last()
            temporal=fac_especialista_detalle.objects.get(id=a.id)
            form = fac_especialista_detalleform(instance=temporal)
        else:
            form = fac_especialista_detalleform()
            
    
    for k in fac_especialista_detalle1:
        contrato_ins=contrato.objects.get(especialista= k.fac_especialista.especialista.id)
        if contrato_ins.reten_art_383 == 'SI':
            k.honorario = total_sum
            k.save()
        else:
            k.honorario = 0
            k.save()

    
    subtotal = float(sum(resul)-sum(c)+ddd.valor   )
    return render(request, 'fac_especialista_detalleform.html',{'form':form,'fac_especialista_detalles':fac_especialista_detalle1,'subtotals':subtotal, 'tempos':tempo})


#vista para el filtro de selecccion
def load_actividad(request):
    centro_costo = request.GET.get('centro_costo')
    centro_actividades = centro_actividad.objects.filter(centro_costo=centro_costo).order_by('actividad')
    return render(request,'lista_Activiades.html',{'centro_actividades':centro_actividades})


# ESTA CLASE ES PARA listar La toda fcaturacion detallada y esta OCULTA
def fac_especialista_detalle_list(request):
    fac_especialista_detalle1= fac_especialista_detalle.objects.all()
    contexto = {'fac_especialista_detalles': fac_especialista_detalle1}
    return render(request,'fac_especialista_detalle_list.html', contexto) 


##################==NOTA: esta vista de lista no la ve el usuario, esta oculta==###########################
class fac_especialista_detalleEdit(UpdateView):
    model=fac_especialista_detalle
    form_class=fac_especialista_detalleform
    template_name='fac_especialista_detalleform.html'
    success_url=reverse_lazy('fac_especialista_detalle_list')
#[[[[[[[[[====FINALIZA FAC_ESPECIALISTAS_DETALLE=========[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]
#**************************************************************************************************************************************************************************




#[[[[[[[[[========INICIA FAC_ESPECIALISTAS========[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]
#**************************************************************************************************************************************************************************
#ESTA CLASE ES PARA CREAR LA FACTURA DE LOS ESPECIALISTAS
class fac_especialistaCrear(CreateView):
    model = fac_especialista
    form_class = fac_especialistaform
    template_name = 'fac_especialista_form.html'
    success_url = reverse_lazy('fac_especialista_view')


#ESTA FUNCION ES PARA CREAR UNA LISTA DE LOS FAC_ESPECIALISTAS
def fac_especialista_list(request):

    sum_general=fac_especialista.objects.all()
    cont_sum_general = fac_especialista.objects.count()
    conttt = 1
    for i in sum_general:
        #esto es para calcular el honorario del mes actual
        aaa = fac_especialista_detalle.objects.filter(fac_especialista=i.id)
        bbb = tipo_fact.objects.all()
        ccc = centro_actividad.objects.all()
        contr = contrato.objects.get(especialista = i.especialista)
        ddd = contrato.objects.get(especialista = i.especialista)
        suma_parcial=factura_pdf(aaa, bbb, ccc, ddd)
        (resul,total_sum, a, vec_ev, c, d,e, f, g)=suma_parcial.opera_especialistas()
        i.valor = sum(resul) - sum(c) + contr.valor-sum(f)
        i.save()
        
        #esto es para calcular el moto acumulado del especialista hasta la ultima factura
        vector = fac_especialista.objects.filter(especialista= i.especialista)
        vect_ult = vector.last()
        cont_vector = vector.count()
        sum_acum = 0

        if vect_ult.id == i.id:
            for j in vector:
                sum_acum += j.valor
            i.acum = sum_acum
        
        i.save() 
     
             
        
    filtrofechacpp=request.GET.get("filtrofechacpp")#extrae de especialistas_list.html el valor de la entrada de texto llamada "filtroname" que corresponde al nombre del especialista

    filtroidcpp=request.GET.get("filtroidcpp")#extrae de especialistas_list.html el valor de la entrada de texto llamada "filtroid" que corresponde al nombre del especialista
    def __str__(self):
        return '%s ' % (self.filtroidcpp)
    

    if filtrofechacpp:
        fac_especialista1 = fac_especialista.objects.filter(fechafac_esp__contains = filtrofechacpp)#aqui se filtran los objetos que cumplen el filtr de nombre
       
        if filtroidcpp:#aqui se filtra nuevamente pero por "id" a los especialistas que cumplieron con el filtro "nombre"
            fac_especialista1 = fac_especialista.objects.filter(id__icontains=filtroidcpp)       
        
        contexto = {'fac_especialistas':fac_especialista1}
        return render(request,'fac_especialista_list.html', contexto)
    else:#este else se hace para que el usuario pueda filtrar por cualquiera de los dos filtros ademas de los dos al mismo tiempo

        if filtroidcpp:
            fac_especialista1 = fac_especialista.objects.filter(id__icontains= filtroidcpp)#primero se filtra por "id"
        
            if filtrofechacpp:#despues se filtra por "nombre"
                fac_especialista1 = fac_especialista.objects.filter(fechafac_esp__contains=filtrofechacpp)       
        
            contexto = {'fac_especialistas':fac_especialista1}
            return render(request,'fac_especialista_list.html', contexto)
        else:#si los filtros estan en blanco se muestran todos los registros
            
            fac_especialista1 = fac_especialista.objects.all()
            contexto = {'fac_especialistas':fac_especialista1}
            return render(request,'fac_especialista_list.html', contexto)


class retencionCrear(CreateView):
    model = retencion
    form_class = retencionform
    template_name = 'especialistas/retencion_form.html'
    success_url = reverse_lazy('retencion_list')
    
def retencion_list(request, id_):
    obj2 = fac_especialista.objects.get(id = id_)
    retencion1= retencion.objects.filter(fac_especialista = id_)
    obj1 = contrato.objects.get(especialista = obj2.especialista)
    obj3 = uvt.objects.get(id = obj2.tarifa.id)
    obj4 = reten_383.objects.filter(tarifa = obj2.tarifa.id)
    calculo = factura_pdf(obj1, obj2, obj3, obj4)

    
    (honorario_cal,
     incr_aport_pension_cal, 
     incr_solida_pensional_cal,
     incr_aport_salud_cal, 
     incr_aport_arl_cal, 
     incr_aport_vol_pension_cal,
     aport_volun_empleador_Cal,
     indemni_lab_cal, 
     re_rent_exent_lab_cal, 
     re_deduc_rent_exent_cal, 
     re_tope_rent_exent_lab_cal, 
     re_total_base_grav_reten_cal,
     re_valor_reten_cal,
     deduc_int_prest_vivienda_cal, 
     deduc_plan_comp_salud_cal, 
     deduc_depen_cargo_cal, 
     re_base_grav_reten_uvt_cal, 
     re_fuente_uvt_cal,
     aport_cuenta_afc_cal,
     
     total_ingre_no_rent_cal,
     total_deducciones_cal,
     total_rent_exten_dos_cal
     
     ) = calculo.retencion_esp()
    
    
    cont_retencion1 = retencion1.count()
    
    if cont_retencion1 == 0:
        retencion_prov = retencion()
        
        retencion_prov.fac_especialista = id_
        retencion_prov.name_especialista = obj2.especialista.name_esp
        retencion_prov.honorario = honorario_cal
        
        retencion_prov.incr_aport_pension = incr_aport_pension_cal
        retencion_prov.incr_solida_pensional = incr_solida_pensional_cal
        retencion_prov.incr_aport_salud =incr_aport_salud_cal
        retencion_prov.incr_aport_arl = incr_aport_arl_cal
        retencion_prov.incr_aport_vol_pension  = incr_aport_vol_pension_cal
        
        retencion_prov.deduc_int_prest_vivienda = deduc_int_prest_vivienda_cal
        retencion_prov.deduc_plan_comp_salud = deduc_plan_comp_salud_cal
        retencion_prov.deduc_depen_cargo = deduc_depen_cargo_cal
        
        retencion_prov.aport_cuenta_afc = aport_cuenta_afc_cal
        retencion_prov.aport_volun_empleador = aport_volun_empleador_Cal
        retencion_prov.indemni_lab = indemni_lab_cal
       
        retencion_prov.re_rent_exent_lab = re_rent_exent_lab_cal
        retencion_prov.re_deduc_rent_exent = re_deduc_rent_exent_cal
        retencion_prov.re_tope_rent_exent_lab = re_tope_rent_exent_lab_cal
        retencion_prov.re_total_base_grav_reten = re_total_base_grav_reten_cal
        retencion_prov.re_base_grav_reten_uvt = re_base_grav_reten_uvt_cal
        retencion_prov.re_fuente_uvt = re_fuente_uvt_cal
        retencion_prov.re_valor_reten = re_valor_reten_cal
        
        retencion_prov.save()
    else:
        elim_prov = retencion.objects.get(fac_especialista = id_)
        elim_prov.delete()
        
        retencion_prov = retencion()
        
        retencion_prov.fac_especialista = id_
        retencion_prov.name_especialista = obj2.especialista.name_esp
        retencion_prov.honorario = honorario_cal
        
        retencion_prov.incr_aport_pension = incr_aport_pension_cal
        retencion_prov.incr_solida_pensional = incr_solida_pensional_cal
        retencion_prov.incr_aport_salud =incr_aport_salud_cal
        retencion_prov.incr_aport_arl = incr_aport_arl_cal
        retencion_prov.incr_aport_vol_pension  = incr_aport_vol_pension_cal
        
        retencion_prov.deduc_int_prest_vivienda = deduc_int_prest_vivienda_cal
        retencion_prov.deduc_plan_comp_salud = deduc_plan_comp_salud_cal
        retencion_prov.deduc_depen_cargo = deduc_depen_cargo_cal
        
        retencion_prov.aport_cuenta_afc = aport_cuenta_afc_cal
        retencion_prov.aport_volun_empleador = aport_volun_empleador_Cal
        retencion_prov.indemni_lab = indemni_lab_cal
       
        retencion_prov.re_rent_exent_lab = re_rent_exent_lab_cal
        retencion_prov.re_deduc_rent_exent = re_deduc_rent_exent_cal
        retencion_prov.re_tope_rent_exent_lab = re_tope_rent_exent_lab_cal
        retencion_prov.re_total_base_grav_reten = re_total_base_grav_reten_cal
        retencion_prov.re_base_grav_reten_uvt = re_base_grav_reten_uvt_cal
        retencion_prov.re_fuente_uvt = re_fuente_uvt_cal
        retencion_prov.re_valor_reten = re_valor_reten_cal
        
        retencion_prov.save()
        

    contexto = {'retencions': retencion1}
    return render(request, 'especialistas/retencion_list.html', contexto)
       

#[[[[[[[[[====FINALIZA FAC_ESPECIALISTAS============[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]
#**************************************************************************************************************************************************************************




#[[[[[[[[[=======INICIA ESPECIALISTA_CPP_AUX===========[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]
#**************************************************************************************************************************************************************************
#ESTA CALSE ES PARA CREAR LA tabla auxiliar DE LOS ESPECIALISTAS
class especialista_cpp_auxCrear(CreateView):
    model = especialista_cpp_aux
    form_class = especialista_cpp_auxform
    template_name = 'especialista_cpp_aux_form.html'
    success_url = reverse_lazy('especialista_cpp_aux_list')


class especialista_cpp_auxEdit(UpdateView):
    model = especialista_cpp_aux
    form_class = especialista_cpp_auxform
    template_name = 'especialista_cpp_aux_form.html'
    success_url = reverse_lazy('especialista_cpp_aux_list')


def especialista_cpp_aux_list(request):
    especialista_cpp_aux1= especialista_cpp_aux.objects.all()
    contexto = {'especialista_cpp_auxs': especialista_cpp_aux1}
    return render(request, 'especialista_cpp_aux_list.html', contexto)
#[[[[[[[[[====FINALIZA ESPECILIASTA_CPP_AUX========[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]
#**************************************************************************************************************************************************************************





#[[[[[[[[[====INICIO DE CENTRO_COSTO========[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]
#**************************************************************************************************************************************************************************
class centro_costoCrear(CreateView):
    model=centro_costo
    form_class=centro_costoform
    template_name= 'centro_costo_form.html'
    success_url=reverse_lazy('centro_costo_list')


class centro_costoEdit(UpdateView):
    model=centro_costo
    form_class=centro_costoform
    template_name= 'centro_costo_form.html'
    success_url=reverse_lazy('centro_costo_list')


def centro_costo_list(request):
    centro_costo1=centro_costo.objects.all()
    contexto={'centro_costos': centro_costo1}
    return render(request,'centro_costo_list.html',contexto)
#[[[[[[[[[====FINALIZA DE CENTRO_COSTO========[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]
#**************************************************************************************************************************************************************************




#[[[[[[[[[====INICIO DE ACTIVIDAD========[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]
#**************************************************************************************************************************************************************************
class actividadCrear(CreateView):
    model=actividad
    form_class=actividadform
    template_name='actividad_form.html'
    success_url=reverse_lazy('actividad_list')


class actividadEdit(UpdateView):
    model=actividad
    form_class=actividadform
    template_name='actividad_form.html'
    success_url=reverse_lazy('actividad_list')


def actividad_list(request):
    actividad1=actividad.objects.all()
    contexto={'actividads':actividad1}
    return render(request,'actividad_list.html',contexto) 
#[[[[[[[[[====FINALIZA DE ACTIVIDAD========[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]
#**************************************************************************************************************************************************************************




#[[[[[[[[[====INICIA SUB_ACTIVITY========[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]
#**************************************************************************************************************************************************************************
class sub_activityCrear(CreateView):
    model= sub_activity
    form_class= sub_activityform
    template_name='sub_activity_form.html'
    success_url=reverse_lazy('sub_activity_list')


class sub_activityEdit(UpdateView):
    model= sub_activity
    form_class= sub_activityform
    template_name='sub_activity_form.html'
    success_url = reverse_lazy('sub_activity_list')


def sub_activity_list(request):
    sub_activity1=sub_activity.objects.all()
    contexto={'sub_activitys':sub_activity1}
    return render(request,'sub_activity_list.html',contexto) 
#[[[[[[[[[====FINALIZA SUB_ACTIVITY========[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]
#**************************************************************************************************************************************************************************






#[[[[[[[[[====INICIA FUNCION========[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]
#**************************************************************************************************************************************************************************
#ESTA ES LA LISTA QUE MUESTRA LOS DETALLES DE LA FATURA HECHA EN FAC_ESPECIALISTA
def detalle_facturador(request, id_):

    filtro_Esp = fac_especialista_detalle.objects.filter(fac_especialista=id_)
    suc_actt = sub_activity.objects.all()
    cc = centro_costo.objects.all()
    tf= tipo_fact.objects.all()
    resp=factura_pdf(filtro_Esp, cc ,suc_actt,tf)
    resp.factura_esp()

    
    fac_especialista_detalle1=fac_especialista_detalle.objects.filter(fac_especialista=id_)
    contexto={'fac_especialista_detalles':fac_especialista_detalle1}
    return render(request,'index.html', contexto) 

#EDITA EL DETALLE DE LOS CPP  O  FATURAS
def fac_especialista_detalle_detalEdit(request, id_):
    fac_especialista_detalleEdit1 = fac_especialista_detalle.objects.get(id=id_)
    if request.method == 'GET':
        form = fac_especialista_detalleform(instance = fac_especialista_detalleEdit1)
    else:
        form = fac_especialista_detalleform(request.POST, instance = fac_especialista_detalleEdit1)
        if form.is_valid():
            form.save()
        return redirect('fac_especialista_list')
    contexto={'form':form}
    return render(request, 'index_form.html', contexto) 


#[[[[[[[[[====FINALIZA FUNCION========[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[[]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]]
#**************************************************************************************************************************************************************************


#aqui van los tipos de tarifa
def tarifa_list(request):
    tarifa1 = Tarifa.objects.all()
    contexto = {'tarifas':tarifa1}
    return render(request,'especialistas/tarifa_list.html', contexto)

class tarifaCrear(CreateView):
    model = Tarifa
    form_class = Tarifaform
    template_name= 'especialistas/tarifa_form.html'
    success_url = reverse_lazy('tarifa_list')

class tarifaEdit(UpdateView):
    model = Tarifa
    form_class = Tarifaform
    template_name= 'especialistas/tarifa_form.html'
    success_url = reverse_lazy('tarifa_list')

#=====================================================================================================================================================================================================================================
#|||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||
#|||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||
#|||||||||||||=====AQUI VAN LOS SERVICIOS MEDICOS======||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||
#|||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||
#|||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||
#=====================================================================================================================================================================================================================================
#LISTAR SERVICIOS MEDICOS
def servicio_medico_list(request):
    servicio_medico1=servicio_medico.objects.all()
    contexto={'servicio_medicos':servicio_medico1}
    return render(request,'servicios_medicos/servicio_medico_list.html',contexto) 

#CREAR O REGISTRAR SERVICIO MEDICO
class  servicio_medicoCrear(CreateView):
    model =servicio_medico
    form_class= servicio_medicoform
    template_name= 'servicios_medicos/servicio_medico_form.html'
    success_url=reverse_lazy('servicio_medico_list')

#EDITAR SERVICO MEDICO
class servicio_medicoEdit(UpdateView):
    model =servicio_medico
    form_class= servicio_medicoform
    template_name= 'servicios_medicos/servicio_medico_form.html'
    success_url=reverse_lazy('servicio_medico_list')
#_________________________________________________________________________________________________

def cpp_servicio_medico_list(request):

    sum_general1 = cpp_servicio_medico.objects.all()
    for j in sum_general1:
        temp=cpp_servicio_medico_detalle.objects.filter(cpp_servicio_medico=j.id)
        ee=factura_pdf(temp)
        j.valor_sm=ee.summ_sm()
        j.save()


    cpp_servicio_medico1=cpp_servicio_medico.objects.all()
    contexto={'cpp_servicio_medicos':cpp_servicio_medico1}
    return render(request,'servicios_medicos/cpp_servicio_medico_list.html',contexto) 

class cpp_servicio_medicoCrear(CreateView):
    model = cpp_servicio_medico
    form_class = cpp_servicio_medicoform
    template_name = 'servicios_medicos/cpp_servicio_medico_form.html'
    success_url = reverse_lazy('cpp_servicio_medico_list')

#___________________________________________________________________________________________________________________

#lista las facturas registradas
def cpp_servicio_medico_detalle_list(request):
    cpp_servicio_medico_detalle1=cpp_servicio_medico_detalle.objects.all()
    contexto={'cpp_servicio_medico_detalles':cpp_servicio_medico_detalle1}
    return render(request, 'servicios_medicos/cpp_servicio_medico_detalle_list.html',contexto)


def cpp_servicio_medico_detalleCrear(request):
    
    contar_detalle_sm=cpp_servicio_medico_detalle.objects.count()

    filtro=cpp_servicio_medico_detalle.objects.all()
    
    for i in filtro:
        a = i.id

    if request.method == 'POST' :
        form = cpp_servicio_medico_detalleform(request.POST)
        if form.is_valid(): 
            form.save()
        return redirect('cpp_servicio_medico_detalle_view')
    else:
        if contar_detalle_sm != 0:
            temporal=cpp_servicio_medico_detalle.objects.get(id=a)
            form = cpp_servicio_medico_detalleform(instance=temporal)
        else:
            form = cpp_servicio_medico_detalleform()
    return render(request, 'servicios_medicos/cpp_servicio_medico_detalle_form.html', {'form':form})



#esta clase edita a cada uno de los elementos registrados 
class cpp_servicio_medico_detalleEdit(UpdateView):
    model=cpp_servicio_medico_detalle
    form_class = cpp_servicio_medico_detalleform
    template_name= 'servicios_medicos/cpp_servicio_medico_detalle_form.html'
    success_url= reverse_lazy('cpp_servicio_medico_detalle_list')
#________________hacer la lista que muestra lows detalles______________________________________________________________________________________________


#esta es la fuuncion que edita el detalle  de cada una de las facturas
def cpp_servicio_medico_detalleEditdetal(request, id_):
    cpp_servicio_medico_detalle1 = cpp_servicio_medico_detalle.objects.get(id=id_)
    if request.method == 'GET':
        form = cpp_servicio_medico_detalleform(instance = cpp_servicio_medico_detalle1)
    else:
        form = cpp_servicio_medico_detalleform(request.POST, instance = cpp_servicio_medico_detalle1)
        if form.is_valid():
            form.save()
        return redirect('cpp_servicio_medico_list')
    contexto={'form':form}
    return render(request, 'servicios_medicos/cpp_servicio_medico_detalle_form.html', contexto) 


def cpp_servicio_medico_detalle_listadetal(request, id_):
    cpp_servicio_medico_detalle1 = cpp_servicio_medico_detalle.objects.filter(cpp_servicio_medico=id_)
    contexto={'cpp_servicio_medico_detalles':cpp_servicio_medico_detalle1}
    return render(request, 'servicios_medicos/cpp_sm_detal_list.html',contexto)







#=====================================================================================================================================================================================================================================
#|||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||
#|||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||
#|||||||||||||=====AQUI VAN LOS ARRIENDOS======||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||
#|||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||
#|||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||
#=====================================================================================================================================================================================================================================

#cuentas auxiliares
def cuenta_arriendo_aux_list(request):
    cuenta_aux_arri = cuenta_arriendo_aux.objects.all()
    contexto = {'cuenta_arriendo_auxs':cuenta_aux_arri}
    return render(request, 'arriendos/cuenta_arriendo_aux_list.html', contexto)


class cuenta_arriendo_auxCrear(CreateView):
    model = cuenta_arriendo_aux
    form_class = cuenta_arriendo_auxform
    template_name = 'arriendos/cuenta_arriendo_aux_form.html'
    success_url = reverse_lazy('cuenta_arriendo_aux_list')
    
class cuenta_arriendo_auxEdit(UpdateView):
    model = cuenta_arriendo_aux
    form_class = cuenta_arriendo_auxform
    template_name = 'arriendos/cuenta_arriendo_aux_form.html'
    success_url = reverse_lazy('cuenta_arriendo_aux_list')   
#_____________________________________________________________


#listar los arriendos registrados
def arriendo_list(request):
    arriendo1=arriendo.objects.all()
    contexto={'arriendos':arriendo1}
    return render(request, 'arriendos/arriendo_list.html',contexto)

class arriendoCrear(CreateView):
    model = arriendo
    form_class = arriendoform
    template_name = 'arriendos/arriendo_form.html'
    success_url = reverse_lazy('arriendo_list')

class arriendoEdit(UpdateView):
    model = arriendo
    form_class = arriendoform
    template_name = 'arriendos/arriendo_form.html'
    success_url = reverse_lazy('arriendo_list')
#______________________________________________________________

#cpp arriendos
def cpp_arriendo_list(request):
    
    filtroobj=cpp_arriendo.objects.all()
    for i in filtroobj:
        obj1=cpp_arriendo.objects.get(id=i.id)
        obj2=inductor_arri.objects.filter(arriendo=obj1.arriendo)
        inductor_arriendo_cpp = factura_pdf(obj1,obj2)
        vector_resultado = inductor_arriendo_cpp.calculo_cpp_arriendo() 
        sumaT = inductor_arriendo_cpp.suma_cpp_arriendo() 
        
        i.valor_cont=sumaT
        i.save()
        
    cpp_arriendo1= cpp_arriendo.objects.all()
    contexto={'cpp_arriendos':cpp_arriendo1}
    return render(request,'arriendos/cpp_arriendo_list.html',contexto)


class cpp_arriendoCrear(CreateView):
    model = cpp_arriendo
    form_class = cpp_arriendoform
    template_name = 'arriendos/cpp_arriendo_form.html'
    success_url = reverse_lazy('cpp_arriendo_list')
#_____________________________________________________________________

#cpp arriendos detalle(esta vista e3sta oculta)
def cpp_arriendo_detalle_list(request):
    cpp_arriendo_detalle1 = cpp_arriendo_detalle.objects.all()
    contexto={'cpp_arriendo_detalles':cpp_arriendo_detalle1}
    return render(request, 'arriendos/cpp_arriendo_detalle_list.html', contexto)

#esta clase edita el detalle que se encuentra oculta
class cpp_arriendo_detalleEdit(UpdateView):
    model = cpp_arriendo_detalle
    form_class = cpp_arriendo_detalleform
    template_name = 'arriendos/cpp_arriendo_detalle_form.html'
    success_url = reverse_lazy('cpp_arriendo_detalle_list')
#______________________________________________________________________

#crear cpp detalle de los arriendos[ESTA VISTA NO SE UTILIZA PERO NO SE PUEDE BORRAR porque ocurre un error que no logro arreglar]
def cpp_arriendo_detalleCrear(request):
    
    contar_detalle_arri=cpp_arriendo_detalle.objects.count()

    filtro=cpp_arriendo_detalle.objects.all()
    
    for i in filtro:
        b = i.id

    if request.method == 'POST' :
        form = cpp_arriendo_detalleform(request.POST)
        if form.is_valid(): 
            form.save()
        return redirect('cpp_arriendo_detalle_view')
    else:
        if contar_detalle_arri != 0:
            temporal=cpp_arriendo_detalle.objects.get(id=b)
            form = cpp_arriendo_detalleform(instance=temporal)
        else:
            form = cpp_arriendo_detalleform()
    return render(request, 'arriendos/cpp_arriendo_detalle_form.html', {'form':form})

    #________________________________

#listar  los cpp de arriendos detallados
def cpp_arriendo_detalle_listadetal(request, id_):
    
    i=cpp_arriendo.objects.get(id=id_)
    obj1=cpp_arriendo.objects.get(id=id_)
    obj2=inductor_arri.objects.filter(arriendo=obj1.arriendo)
    inductor_arriendo_cpp = factura_pdf(obj1,obj2)
    vector_resultado = inductor_arriendo_cpp.calculo_cpp_arriendo() 
    cuenta_aux_arri = cuenta_arriendo_aux.objects.filter(arriendo = obj1.arriendo.id)

    cont_induc=obj2.count()
    cont_arri_detal=cpp_arriendo_detalle.objects.filter(cpp_arriendo=id_).count()
    
    if cont_arri_detal == 0:
        k=0
        for j in obj2:
            neww = cpp_arriendo_detalle()
            
            neww.cpp_arriendo = i.id
            neww.name_arri = j.arriendo.name_arri
            # neww.cuenta_reten = obj1.arriendo.cuenta_reten
            neww.centro_costo = j.centro_costo.name_ccos
            neww.inductor_arri= j.induc
            # neww.reten= j.arriendo.reten
            neww.num_cuenta = j.centro_costo.id_centro_costo
            neww.cuenta_especific =j.cuenta_especific
            neww.valor_cpp_arri_detal= round(vector_resultado[k,0])
            neww.fecha_cpp_arri_detal=i.fecha_cpp_arri
            neww.save()
            k += 1    
        
        #registrar las cuentas auxiliares
        r = 0
        for q in cuenta_aux_arri:
            new = cpp_arriendo_detalle()
            
            new.cpp_arriendo = id_
            new.name_arri = q.arriendo.name_arri
            # neww.cuenta_reten = q.cuenta
            new.centro_costo = q.name_cuenta
            
            if q.cuenta[0:4]=="2365":
                new.inductor_arri = obj1.reten
                new.cuenta_especific = q.cuenta
                new.valor_cpp_arri_detal = round(obj1.valor_cpp_arri*obj1.reten/100)
            else:
                new.inductor_arri = 0
                new.cuenta_especific = q.cuenta
                new.valor_cpp_arri_detal = round((obj1.valor_cpp_arri)-(obj1.valor_cpp_arri*obj1.reten/100))

                
            # neww.reten= j.arriendo.reten
            new.num_cuenta = " "
            
            
            new.fecha_cpp_arri_detal= i.fecha_cpp_arri
            
            new.save()
            r += 1
 
    # else:ESTA PARTE ES PARA ACTUALIZAR LOS VALORES DE LOS REGISTROS DEL DETALLE CUANDO SE MODIFIQUEN LOS PARAMETROS DE LOS INDUCTORES
    #     filtro = cpp_arriendo_detalle.objects.filter(cpp_arriendo=i.id)
    #     filtro.delete()
    #     k=0
    #     for j in obj2:
    #         neww = cpp_arriendo_detalle()
    #         neww.cpp_arriendo = i.id
    #         neww.cuenta_reten = obj1.arriendo.cuenta_reten
    #         neww.name_arri = j.arriendo.name_arri
    #         neww.centro_costo = j.centro_costo.name_ccos
    #         neww.inductor_arri= j.induc
    #         neww.reten= j.arriendo.reten
    #         neww.num_cuenta = j.centro_costo.id_centro_costo
    #         neww.cuenta_especific =j.cuenta_especific
    #         neww.valor_cpp_arri_detal=vector_resultado[k,0]
    #         neww.fecha_cpp_arri_detal=i.fecha_cpp_arri
    #         neww.save()
    #         k += 1  
    
    cpp_arriendo_detalle1 = cpp_arriendo_detalle.objects.filter(cpp_arriendo=id_)
    contexto={'cpp_arriendo_detalles':cpp_arriendo_detalle1}
    return render(request, 'arriendos/cpp_arri_detal_list.html',contexto)
    
def load_arriendo(request, id_):#[ESTA VISTA CREA EL DETALLE DE LOS CPP DE ARRIENDO]
    contar_detalle_arri=cpp_arriendo_detalle.objects.count()
    filtro=cpp_arriendo_detalle.objects.all()
    
    for i in filtro:
        c = i.id
    
    if request.method == 'POST' :
        form = cpp_arriendo_detalleform(request.POST)
        if form.is_valid(): 
            form.save()
        return redirect('cpp_arriendo_detalle_view')
    else:
        if contar_detalle_arri != 0:
            intancia_llenar = cpp_arriendo_detalle.objects.get(id=c)
            intancia_llenar.cpp_arriendo = int(id_)
            temporal=intancia_llenar
            form = cpp_arriendo_detalleform(instance=temporal)
        else:
            form = cpp_arriendo_detalleform()
    return render(request, 'arriendos/cpp_arriendo_detalle_form.html', {'form':form})

#editar  los cpp de arriendos detallados
def cpp_arriendo_detalleEditdetal(request, id_):
    cpp_arriendo_detalle1 = cpp_arriendo_detalle.objects.get(id=id_)
    if request.method == 'GET':
        form = cpp_arriendo_detalleform(instance = cpp_arriendo_detalle1)
    else:
        form = cpp_arriendo_detalleform(request.POST, instance = cpp_arriendo_detalle1)
        if form.is_valid():
            form.save()
        return redirect('cpp_arriendo_list')
    contexto={'form':form}
    return render(request, 'arriendos/cpp_arri_detal_form.html', contexto) 
#___________________________________________________________________

#inductor
def inductor_arri_list(request):
    inductor_arri1=inductor_arri.objects.all()
    contexto={'inductor_arris':inductor_arri1}
    return render(request, 'arriendos/inductor_arriendo_list.html', contexto)

class inductor_arriEdit(UpdateView):
    model = inductor_arri
    form_class = inductor_arriform
    template_name = 'arriendos/inductor_arriendo_form.html'
    success_url = reverse_lazy('inductor_arri_list')

def inductor_arriCrear(request):
    
    contar_inductor_arri=inductor_arri.objects.count()

    filtro=inductor_arri.objects.all()
    
    for i in filtro:
        a = i.id

    if request.method == 'POST' :
        form = inductor_arriform(request.POST)
        if form.is_valid(): 
            form.save()
        return redirect('inductor_arri_view')
    else:
        if contar_inductor_arri != 0:
            temporal=inductor_arri.objects.get(id=a)
            form = inductor_arriform(instance=temporal)
        else:
            form = inductor_arriform()
    return render(request, 'arriendos/inductor_arriendo_form.html', {'form':form})


def load_arriendo_ajax(request):
    arriendo = request.GET.get('arriendo')
    inductor_arri2s = inductor_arri.objects.filter(arriendo=arriendo).order_by('arriendo')
    return render(request,'arriendos/lista_arriendo.html',{'inductor_arri2s':inductor_arri2s})






#=====================================================================================================================================================================================================================================
#|||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||
#|||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||
#|||||||||||||=====AQUI VAN LOS PROVEEDORES======||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||
#|||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||
#|||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||
#=====================================================================================================================================================================================================================================

#proveedores registrados
def proveedor_list(request):
    proveedor1 = proveedor.objects.all()
    contexto={'proveedors':proveedor1}
    return render(request, 'proveedores/proveedor_list.html', contexto)

class proveedorCrear(CreateView):
    model = proveedor
    form_class = proveedorform
    template_name = 'proveedores/proveedor_form.html'
    success_url = reverse_lazy('proveedor_list')

class proveedorEdit(UpdateView):
    model = proveedor
    form_class = proveedorform
    template_name = 'proveedores/proveedor_form.html'
    success_url = reverse_lazy('proveedor_list')


def categoria_list(request):
    categoria1 = categoria.objects.all()
    contexto = {'categorias':categoria1}
    return render(request, 'proveedores/categoria_list.html', contexto)


class categoriaEdit(UpdateView):
    model = categoria
    form_class = categoriaform
    template_name = 'proveedores/categoria_form.html'
    success_url = reverse_lazy('categoria_list')

class categoriaCrear(CreateView):
    model = categoria
    form_class = categoriaform
    template_name = 'proveedores/categoria_form.html'
    success_url = reverse_lazy('categoria_list')


def distribucion_list(request):
    distribucion1 = distribucion.objects.all()
    contexto = {'distribucions':distribucion1}
    return render(request, 'proveedores/distribucion_list.html', contexto)

class distribucionEdit(UpdateView):
    model = distribucion
    form_class = distribucionform
    template_name = 'proveedores/distribucion_form.html'
    success_url = reverse_lazy('distribucion_list')


class cuenta_auxCrear(CreateView):
    model = cuenta_aux
    form_class = cuenta_auxform
    template_name = 'proveedores/cuenta_aux_form.html'
    success_url = reverse_lazy('cuenta_aux_list')

class cuenta_auxEdit(UpdateView):
    model = cuenta_aux
    form_class = cuenta_auxform
    template_name = 'proveedores/cuenta_aux_form.html'
    success_url = reverse_lazy('cuenta_aux_list')

def cuenta_aux_list(request):
    cuenta_aux1 = cuenta_aux.objects.all()
    contexto = {'cuenta_auxs':cuenta_aux1}
    return render(request, 'proveedores/cuenta_aux_list.html', contexto)

def distribucionCrear(request):
    
    contar_distri=distribucion.objects.count()
    
    if contar_distri != 0:
        a=distribucion.objects.last().id

    if request.method == 'POST' :
        form = distribucionform(request.POST)
        if form.is_valid(): 
            form.save()
        return redirect('distribucion_view')
    else:
        if contar_distri != 0:
            temporal=distribucion.objects.get(id=a)
            form = distribucionform(instance=temporal)
        else:
            form = distribucionform()
    return render(request, 'proveedores/distribucion_form.html', {'form':form})


class productoCrear(CreateView):
    model = producto
    form_class = productoform
    template_name = 'proveedores/producto_form.html'
    success_url = reverse_lazy('producto_list')

def producto_list(request):
    producto1 = producto.objects.all()
    contexto={'productos':producto1}
    return render(request, 'proveedores/producto_list.html', contexto)

class productoEdit(UpdateView):
    model = producto
    form_class = productoform
    template_name = 'proveedores/producto_form.html'
    success_url = reverse_lazy('producto_list')

# cpp de proveedores
def cpp_proveedor_list(request):
    
    #esta funcion es para calcular el costo total de la cpp
    sum_general = cpp_proveedor.objects.all()
    for i in sum_general:
        aaa=cpp_proveedor_detalle.objects.filter(cpp_proveedor=i.id)
        aa = factura_pdf(aaa)
        i.valor_cpp_prov = aa.summ_prov()
        i.save() 

    cpp_proveedor1 = cpp_proveedor.objects.all()
    contexto={'cpp_proveedors':cpp_proveedor1}
    return render(request, 'proveedores/cpp_proveedor_list.html', contexto)

class cpp_proveedorCrear(CreateView):
    model = cpp_proveedor
    form_class = cpp_proveedorform
    template_name = 'proveedores/cpp_proveedor_form.html'
    success_url = reverse_lazy('cpp_proveedor_list')


#cpp proveedores detallados
def cpp_proveedor_detalle_list(request):
    cpp_proveedor_detalle2 = cpp_proveedor_detalle.objects.all()
    contexto = {'cpp_proveedor_detalles': cpp_proveedor_detalle2}
    return render(request, 'proveedores/cpp_proveedor_detalle_list.html', contexto)

#aqui se captura el detalle del detalle
def cpp_proveedor_subdetalle_list(request, id_):
    
    cc = centro_costo.objects.all()
    cpp_prov = cpp_proveedor.objects.get(id = id_)
    cuenta_auxx = cuenta_aux.objects.filter(proveedor = cpp_prov.proveedor.id)
    distri = distribucion.objects.filter(proveedor = cpp_prov.proveedor.id)
    cpp_prov_detal = cpp_proveedor_detalle.objects.filter(cpp_proveedor = id_)
    cont_cpp_prov_detal = cpp_prov_detal.count()

    categ = categoria.objects.all() 
    cpp_prov_calcular = defprov(cc,distri,cpp_prov_detal,cpp_prov,categ)
    
    (resultado, iva_matriz, distrib_mat, mat_detalle) = cpp_prov_calcular.calcular_cpp_proveedor()
    


    cont_sundetalle = cpp_proveedor_subdetalle.objects.filter(cpp_proveedor = id_).count()
    if cont_sundetalle == 0:#verificar que no se halla llenaod con anterioridad la matriz
        if categ.first().id == distri.last().producto.categoria.id:# se esta analizando un gas 
            contt = 0
            for i in cc:#recorrer el numero de centros de costo almacenados
                if resultado[contt,cont_cpp_prov_detal] != 0:#verificar que solo se ingresen los datos distintos de cero en la matriz "matriz_prov"

                    sub_nuevo = cpp_proveedor_subdetalle()
                    
                    sub_nuevo.cpp_proveedor = cpp_prov.id
                    sub_nuevo.centro_costo = i.id
                    sub_nuevo.proveedor = cpp_prov.proveedor.id
                    sub_nuevo.cant_produc = -1
                    sub_nuevo.cant_flete =  -1
                    sub_nuevo.valor_flete = -1
                    
                    cuenta_povicional = distri.last()
                    
                    sub_nuevo.nombre_cuenta = cuenta_povicional.nombre_cuenta
                    sub_nuevo.cuenta = cuenta_povicional.cuenta
                    sub_nuevo.naturaleza_cuenta = cuenta_povicional.naturaleza_cuenta
                    sub_nuevo.distrib = -1
                    sub_nuevo.name_centro_costo = i.name_ccos
                    sub_nuevo.costo = resultado[contt,cont_cpp_prov_detal]
                    sub_nuevo.cuenta_iva = str(int(iva_matriz[contt,cont_cpp_prov_detal+1]))
                    sub_nuevo.valor_iva = iva_matriz[contt,cont_cpp_prov_detal]
                    sub_nuevo.valor_cuenta_contra =  0               
                    sub_nuevo.fecha_cpp_prov_detal = cpp_prov.fecha_cpp_prov

                    sub_nuevo.save()
                    
                contt += 1   
            
            #aqui se hace el registro de la cuentra contra 
            sub_nu = cpp_proveedor_subdetalle()
            #se calcula la cantidad de plata a retener
            total = sum(resultado[0:cc.count(),cont_cpp_prov_detal])
            can_reten = total*cpp_prov.reten/100
            bandera = 0 #esto es para saber si hay registrada una cuenta retenenora. toma un valor distinto a cero en caso de existir     
            
            #Esto es para verificar si existe una cuenta de Retencion antes de iniciar los registros
            for r in cuenta_auxx:
                if r.cuenta[0:4]== "2365":
                    bandera += 1
            
            for m in cuenta_auxx:
                
                if m.cuenta[0:4]== "2365":#esto es una retencion
                    
                    sub_nu.cpp_proveedor = cpp_prov.id
                    sub_nu.centro_costo = -1
                    sub_nu.proveedor = cpp_prov.proveedor.id
                    sub_nu.cant_produc = -1
                    sub_nu.cant_flete =  -1
                    sub_nu.valor_flete = -1
                    sub_nu.distrib = -1
                    sub_nu.name_centro_costo = " "
                    sub_nu.costo = 0
                    sub_nu.cuenta_iva = " "
                    sub_nu.valor_iva = 0
            
                    sub_nu.naturaleza_cuenta = m.naturaleza_cuenta
                    sub_nu.nombre_cuenta = m.name_cuenta
                    sub_nu.cuenta = m.cuenta
                    sub_nu.valor_cuenta_contra =  can_reten
                
                    sub_nu.fecha_cpp_prov_detal = cpp_prov.fecha_cpp_prov  
                    
                    sub_nu.save()        
                else:#no es una cuenta de retencion
                    
                    if bandera != 0:
                        can_reten = can_reten
                    else:
                        can_reten = 0
                      
                    #========================================================
                    sub_nu.cpp_proveedor = cpp_prov.id
                    sub_nu.centro_costo = 0
                    sub_nu.proveedor = cpp_prov.proveedor.id
                    sub_nu.cant_produc = -1
                    sub_nu.cant_flete =  -1
                    sub_nu.valor_flete = -1
                    sub_nu.distrib = -1
                    sub_nu.name_centro_costo = " "
                    sub_nu.costo = 0
                    sub_nu.cuenta_iva = " "
                    sub_nu.valor_iva = 0
                    if cpp_prov.aire_medicinal == 'SI':
                        if m.cuenta[0:6]== "233535":#esta cuenta es de Arrendmientos (retenciones)
                            sub_nu.naturaleza_cuenta = m.naturaleza_cuenta
                            sub_nu.nombre_cuenta = m.name_cuenta
                            sub_nu.cuenta = m.cuenta
                            sub_nu.valor_cuenta_contra =  sum(resultado[0:cc.count(),cont_cpp_prov_detal])+sum(iva_matriz[0:cc.count(),cont_cpp_prov_detal]) - can_reten
                    else:
                        if m.cuenta[0:6] == "220505":# estas cuentas son de proveedores
                            sub_nu.naturaleza_cuenta = m.naturaleza_cuenta
                            sub_nu.nombre_cuenta = m.name_cuenta
                            sub_nu.cuenta = m.cuenta
                            sub_nu.valor_cuenta_contra = sum(resultado[0:cc.count(),cont_cpp_prov_detal])+sum(iva_matriz[0:cc.count(),cont_cpp_prov_detal]) - can_reten
                
                    sub_nu.fecha_cpp_prov_detal = cpp_prov.fecha_cpp_prov
            
                    sub_nu.save()     
        else:# se esta analizando un tipo sangre
            
            contt = 0
            for i in cc:#recorrer el numero de centros de costo almacenados
                
                if resultado[contt,0] != 0:#verificar que solo se ingresen los datos distintos de cero en la matriz "matriz_prov"

                    sub_nuevo = cpp_proveedor_subdetalle()
                    
                    sub_nuevo.cpp_proveedor = cpp_prov.id
                    sub_nuevo.centro_costo = i.id
                    sub_nuevo.proveedor = cpp_prov.proveedor.id
                    sub_nuevo.cant_produc = -1
                    sub_nuevo.cant_flete =  -1
                    sub_nuevo.valor_flete = -1
                    
                    cuenta_povicional = distri.last()
                    
                    sub_nuevo.nombre_cuenta = cuenta_povicional.nombre_cuenta
                    sub_nuevo.cuenta = cuenta_povicional.cuenta
                    sub_nuevo.naturaleza_cuenta = cuenta_povicional.naturaleza_cuenta
                    sub_nuevo.distrib = -1
                    sub_nuevo.name_centro_costo = i.name_ccos
                    sub_nuevo.costo = resultado[contt,0]
                    sub_nuevo.cuenta_iva = " "
                    sub_nuevo.valor_iva = iva_matriz[contt,0]
                    sub_nuevo.valor_cuenta_contra = 0
                    sub_nuevo.fecha_cpp_prov_detal = cpp_prov.fecha_cpp_prov

                    sub_nuevo.save()
                    
                contt += 1   
            
            #se calcula la cantidad de plata a retener
            total = sum(resultado[0:cc.count(),0])
            can_reten = total*cpp_prov.reten/100
            
            for ww in cuenta_auxx:
                if ww.cuenta[0:4] == "2365":#esta cuenta es de retencion
                    
                    #aqui se hace el registro de la cuentra contra 
                    sub_nu = cpp_proveedor_subdetalle()
                            
                    sub_nu.cpp_proveedor = cpp_prov.id
                    sub_nu.centro_costo = 0
                    sub_nu.proveedor = cpp_prov.proveedor.id
                    sub_nu.cant_produc = -1
                    sub_nu.cant_flete =  -1
                    sub_nu.valor_flete = -1
                    
                    sub_nu.distrib = -1
                    sub_nu.name_centro_costo = " "
                    sub_nu.costo = 0
                    sub_nu.cuenta_iva = " "
                    sub_nu.valor_iva = 0
                    
                    sub_nu.naturaleza_cuenta = ww.naturaleza_cuenta
                    sub_nu.nombre_cuenta = ww.name_cuenta
                    sub_nu.cuenta = ww.cuenta
                    sub_nu.valor_cuenta_contra = can_reten
                    
                    sub_nu.fecha_cpp_prov_detal = cpp_prov.fecha_cpp_prov
                    
                    sub_nu.save() 
                else:
                    #aqui se hace el registro de la cuentra contra 
                    sub_nu = cpp_proveedor_subdetalle()
                            
                    sub_nu.cpp_proveedor = cpp_prov.id
                    sub_nu.centro_costo = 0
                    sub_nu.proveedor = cpp_prov.proveedor.id
                    sub_nu.cant_produc = -1
                    sub_nu.cant_flete =  -1
                    sub_nu.valor_flete = -1
                    
                    sub_nu.distrib = -1
                    sub_nu.name_centro_costo = " "
                    sub_nu.costo = 0
                    sub_nu.cuenta_iva = " "
                    sub_nu.valor_iva = 0
                
                    sub_nu.naturaleza_cuenta = ww.naturaleza_cuenta
                    sub_nu.nombre_cuenta = ww.name_cuenta
                    sub_nu.cuenta = ww.cuenta
                    sub_nu.valor_cuenta_contra = total - can_reten
                    sub_nu.fecha_cpp_prov_detal = cpp_prov.fecha_cpp_prov
                    
                    sub_nu.save() 

    cpp_proveedor_subdetalle2 = cpp_proveedor_subdetalle.objects.filter(cpp_proveedor = cpp_prov.id)
    contexto = {'cpp_proveedor_subdetalles':cpp_proveedor_subdetalle2}
    return render(request, 'proveedores/cpp_proveedor_subdetalle_list.html', contexto)
    
    
def cpp_proveedor_detalleEdit(request, id_):
    cpp_proveedor_detalle1 = cpp_proveedor_detalle.objects.get(id=id_)
    if request.method == 'GET':
        form = cpp_proveedor_detalleform(instance = cpp_proveedor_detalle1)
    else:
        form = cpp_proveedor_detalleform(request.POST, instance = cpp_proveedor_detalle1)
        if form.is_valid():
            form.save()
        return redirect('/crearcppprovdetal/'+ str(cpp_proveedor_detalle1.cpp_proveedor.id)+'/')
    contexto={'form':form}
    return render(request, 'proveedores/cpp_prov_detal_form.html', contexto)  
    
    

def cpp_proveedor_detalleCrear(request, id_):
    
    categ = categoria.objects.first()
    cpp_prov = cpp_proveedor.objects.get(id = id_)
    cpp_prov_deta = cpp_proveedor_detalle.objects.filter(cpp_proveedor = id_)
    distri = distribucion.objects.filter(proveedor = cpp_prov.proveedor.id).last()
    
    #colocar el precio de cada producto
    if categ.id == distri.producto.categoria.id:
        for i in cpp_prov_deta:
            i.valor_produc = i.producto.precio
            i.save()

    #rellenar detalle en caso de que tengan valores parametrizados y se tenga que añadir manualmente otros rubros
    distri_produc = distribucion.objects.filter(proveedor = cpp_prov.proveedor.id).filter(producto = cpp_prov.producto)
    
    if cpp_prov_deta.count() == 0 and cpp_prov.valor_factura != 0:
        
        detalle_pro = defprov(cpp_prov,distri_produc)
        vector_distribu = detalle_pro.detalle_automatico()

        for i in distri_produc:
            neww = cpp_proveedor_detalle()
            
            neww.cpp_proveedor = cpp_prov
            neww.categoria = categoria.objects.get(id = distri.producto.categoria.id)
            neww.centro_costo = centro_costo.objects.get(id = i.centro_costo.id)
            neww.producto = producto.objects.get(id = i.producto.id)
            neww.cant_produc = cpp_prov.valor_factura*i.distrib/100
            neww.valor_produc = i.producto.precio
            neww.cant_flete = 0
            neww.valor_flete = 0
            neww.fecha_cpp_prov_detal = cpp_prov.fecha_cpp_prov
            neww.save()
    

    # ================================================
    cc = centro_costo.objects.all()
    cuenta_auxx = cuenta_aux.objects.filter(proveedor = cpp_prov.proveedor.id)
    distri = distribucion.objects.filter(proveedor = cpp_prov.proveedor.id)
    cpp_prov_detal = cpp_proveedor_detalle.objects.filter(cpp_proveedor = id_)
    cont_cpp_prov_detal = cpp_prov_detal.count()
    categ = categoria.objects.all() 
    
    cpp_prov_calcular = defprov(cc,distri,cpp_prov_detal,cpp_prov,categ)
    
    (resultado, iva_matriz, distrib_mat, mat_detalle) = cpp_prov_calcular.calcular_cpp_proveedor()
    
    # ================================================

    
    
    
    
    explorar_elem = cpp_proveedor_detalle.objects.filter(cpp_proveedor = id_)
    contar_elem = explorar_elem.count()
    cp_prov = cpp_proveedor.objects.get(id =id_)
    
    
    categoria_prin = categoria.objects.first()#capturar el primer objeto de la tabla categoria
    cat_prin = categoria_prin.id #capturar el id del primer elemento del objeto categoria
    
    
    distrib = distribucion.objects.filter(proveedor = cp_prov.proveedor).last()
    
    cat_cpp_0 = distrib.producto.categoria.id
    

    if contar_elem != 0:#esto se hace porque inicialmente el cpp_proveedor_detalle esta vacio y manda un error en la plantilla
        categoria_cpp = explorar_elem.last()#captura el ultimo objeto de la tabla cpp_proveedor_detalle
        cat_cpp = categoria_cpp.producto.categoria.id#captura el id del ultimo elemnto del objeto cpp_proveedor_detalle
    else:
        if cat_prin == cat_cpp_0:
            cat_cpp = cat_cpp_0
        else:
            cat_cpp = 0
              
    if request.method == 'POST':
        form = cpp_proveedor_detalleform(request.POST)
        if form.is_valid(): 
            form.save()
        return redirect('/crearcppprovdetal/'+str(id_)+'/')
    else:
        if contar_elem != 0:
            a=explorar_elem.last()
            temporal=cpp_proveedor_detalle.objects.get(id=a.id)
            form = cpp_proveedor_detalleform(instance=temporal)
        else:
            form = cpp_proveedor_detalleform()
    return render(request, 'proveedores/cpp_proveedor_detalle_form.html', {'form':form, 'cpp_proveedor_detalles':explorar_elem, 'id_':id_, 'cat_prins':cat_prin, 'cat_cpps':cat_cpp, 'cp_provs':cp_prov})



#esta vista es para el boton que sirve para observar los detalles de cpp provedores
def cpp_prov_detal_list(request, id_):
    cpp_proveedor_detalle1 = cpp_proveedor_detalle.objects.filter(cpp_proveedor=id_)
    contexto={'cpp_proveedor_detalles':cpp_proveedor_detalle1}
    return render(request, 'proveedores/cpp_prov_detal_list.html',contexto)


def cpp_prov_detalEdit(request, id_):
    cpp_proveedor_detalle1 = cpp_proveedor_detalle.objects.get(id=id_)
    if request.method == 'GET':
        form = cpp_proveedor_detalleform(instance = cpp_proveedor_detalle1)
    else:
        form = cpp_proveedor_detalleform(request.POST, instance = cpp_proveedor_detalle1)
        if form.is_valid():
            form.save()
        return redirect('cpp_proveedor_list')
    contexto={'form':form}
    return render(request, 'proveedores/cpp_prov_detal_form.html', contexto) 


#vista para el filtro de selecccion
def load_producto(request):
    categoria = request.GET.get('categoria')
    producto_1 = producto.objects.filter(categoria=categoria).order_by('nombre')
    return render(request,'proveedores/load_producto.html',{'productos':producto_1})


#centro ce costo desde cpp proveddor detalle 
def load_cpp_prov_detalle(request):
    producto = request.GET.get('producto')
    filtro = distribucion.objects.filter(producto=producto)
    centro_costo_copia = copy.copy(centro_costo.objects.all())
    for i in centro_costo_copia:
        cont = 0
        for j in filtro:
            if i.id == j.centro_costo.id:
                cont += 1
        if cont == 0:
            i.delete()
    return render(request,'proveedores/load_centro_costo.html',{'centro_costos':centro_costo_copia})


#=====================================================================================================================================================================================================================================
#|||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||
#|||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||
#|||||||||||||=====AQUI VAN LOS SERVICIOS======||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||
#|||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||
#|||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||
#=====================================================================================================================================================================================================================================

#estas vistas corresponden a la  creacion o registro de los terceros de los servicios publicos
class serv_publicCrear(CreateView):
    model = serv_public
    form_class = serv_publicform
    template_name = 'servicios_publicos/serv_public_form.html'
    success_url = reverse_lazy('serv_public_list')

class serv_publicEdit(UpdateView):
    model = serv_public
    form_class = serv_publicform
    template_name = 'servicios_publicos/serv_public_form.html'
    success_url = reverse_lazy('serv_public_list')

def serv_public_list(request):
    serv_public1 = serv_public.objects.all()
    contexto={'serv_publics':serv_public1}
    return render(request, 'servicios_publicos/serv_public_list.html',contexto)


#Estas son las vista para los tipos de servicio (publicos o medicos)
class tipo_servCrear(CreateView):
    model = tipo_serv
    form_class = tipo_servform
    template_name = 'servicios_publicos/tipo_serv_form.html'
    success_url = reverse_lazy('tipo_serv_list')

class tipo_servEdit(UpdateView):
    model = tipo_serv
    form_class = tipo_servform
    template_name = 'servicios_publicos/tipo_serv_form.html'
    success_url = reverse_lazy('tipo_serv_list')

def tipo_serv_list(request):
    tipo_serv1 = tipo_serv.objects.all()
    contexto={'tipo_servs':tipo_serv1}
    return render(request, 'servicios_publicos/tipo_serv_list.html',contexto)


#estas vistas corresponden a la  creacion o registro de las cpp de los servicios publicos
class cpp_serv_publicCrear(CreateView):
    model = cpp_serv_public
    form_class = cpp_serv_publicform
    template_name = 'servicios_publicos/cpp_serv_public_form.html'
    success_url = reverse_lazy('cpp_serv_public_list')
    
class cpp_serv_publicEdit(UpdateView):
    model = cpp_serv_public
    form_class = cpp_serv_publicform
    template_name = 'servicios_publicos/cpp_serv_public_form.html'
    success_url = reverse_lazy('cpp_serv_public_list')
    
def cpp_serv_public_list(request):
    
    #hallar el costo total de la factura
    for i in cpp_serv_public.objects.all():
        i.total = i.costo*(1+(i.iva/100))
        i.save()

    
    cpp_serv_public1 = cpp_serv_public.objects.all()
    contexto={'cpp_serv_publics':cpp_serv_public1}
    return render(request, 'servicios_publicos/cpp_serv_public_list.html',contexto)



#estas vistas corresponden a la  creacion o registro de las cuenta_aux_serv
class cuenta_aux_servCrear(CreateView):
    model = cuenta_aux_serv
    form_class = cuenta_aux_servform
    template_name = 'servicios_publicos/cuenta_aux_serv_form.html'
    success_url = reverse_lazy('cuenta_aux_serv_list')
    
class cuenta_aux_servEdit(UpdateView):
    model = cuenta_aux_serv
    form_class = cuenta_aux_servform
    template_name = 'servicios_publicos/cuenta_aux_serv_form.html'
    success_url = reverse_lazy('cuenta_aux_serv_list')
    
def cuenta_aux_serv_list(request):
    cuenta_aux_serv1 = cuenta_aux_serv.objects.all()
    contexto={'cuenta_aux_servs':cuenta_aux_serv1}
    return render(request, 'servicios_publicos/cuenta_aux_serv_list.html',contexto)



#estas vistas corresponden a la  creacion o registro de las distri_serv_public
class distri_serv_publicCrear(CreateView):
    model = distri_serv_public
    form_class = distri_serv_publicform
    template_name = 'servicios_publicos/distri_serv_public_form.html'
    success_url = reverse_lazy('distri_serv_public_list')
    
def distri_serv_publicCrear(request):
    ultimo_distri_serv_public = distri_serv_public.objects.last()
    cont = distri_serv_public.objects.all().count()
    if request.method == 'POST' :
        form = distri_serv_publicform(request.POST)
        if form.is_valid(): 
            form.save()
        return redirect('distri_serv_public_view')
    else:
        if cont != 0 :
            form = distri_serv_publicform(instance=ultimo_distri_serv_public)
        else:
            form = distri_serv_publicform()
    return render(request, 'servicios_publicos/distri_serv_public_form.html', {'form':form})


class distri_serv_publicEdit(UpdateView):
    model = distri_serv_public
    form_class = distri_serv_publicform
    template_name = 'servicios_publicos/distri_serv_public_form.html'
    success_url = reverse_lazy('distri_serv_public_list')
    
def distri_serv_public_list(request):
    distri_serv_public1 = distri_serv_public.objects.all()
    contexto={'distri_serv_publics':distri_serv_public1}
    return render(request, 'servicios_publicos/distri_serv_public_list.html',contexto)

#estas vistas corresponden a la  creacion o registro de las cpp_servi_detalle
    
    
def cpp_servi_detalleEdit(request, id_):
    
    cpp_servi_detalle1 = cpp_servi_detalle.objects.get(id=id_)
    if request.method == 'GET':
        form = cpp_servi_detalleform(instance = cpp_servi_detalle1)
    else:
        form = cpp_servi_detalleform(request.POST, instance = cpp_servi_detalle1)
        if form.is_valid():
            form.save()
        return redirect('/listcpp_servi_detalle/'+str(cpp_servi_detalle1.cpp_serv_public)+'/')
    contexto={'form':form}
    return render(request, 'servicios_publicos/cpp_servi_detalle_form.html', contexto)     


    
def cpp_servi_detalle_list(request, id_):
    
    cpp_ser_pub = cpp_serv_public.objects.get(id = id_)
    distribu = distri_serv_public.objects.filter(serv_public = cpp_ser_pub.serv_public.id)
    
    calcular_distri = funcion_serv(cpp_ser_pub,distribu)
    vector_distri = calcular_distri.distri_serv()
    
    cont_serv_detalle = cpp_servi_detalle.objects.filter(cpp_serv_public = id_).count()
    if cont_serv_detalle == 0:
        i = 0
        for j in distribu:
            neww = cpp_servi_detalle()
            
            neww.cpp_serv_public = id_
            neww.id_serv_public = cpp_ser_pub.serv_public.id
            neww.name_serv_public = cpp_ser_pub.serv_public.nombre_tercero
            neww.id_centro_costo = j.centro_costo.id 
            neww.name_ccos = j.centro_costo.name_ccos
            neww.distri = j.distri
            neww.name_cuenta_aux = j.name_cuenta_especific
            neww.num_cuenta_aux = j.num_cuenta_especific
            neww.costo = vector_distri[i,0]
            neww.num_cuenta_iva = j.num_cuenta_iva
            neww.valor_iva = vector_distri[i,1]
            neww.valor_contra = 0
            neww.fecha = cpp_ser_pub.fecha
            
            neww.save()
            
            i += 1
        
        #obtener las cuentas auxiliares del servicio actual
        serv_auxiliar = cuenta_aux_serv.objects.filter(serv_public=cpp_ser_pub.serv_public.id) 
        
        #contar cuantas cuentas hay alamacenadas
        contar_serv_auxiliar = serv_auxiliar.count() 
        
        if contar_serv_auxiliar != 0:
            ii = 0
            for k in serv_auxiliar:
                
                nuevo = cpp_servi_detalle()
            
                nuevo.cpp_serv_public = id_
                nuevo.id_serv_public = cpp_ser_pub.serv_public.id
                nuevo.name_serv_public = cpp_ser_pub.serv_public.nombre_tercero
                nuevo.id_centro_costo = int(0)
                nuevo.name_ccos = " "
                nuevo.distri = int(0)
                nuevo.name_cuenta_aux = k.name_cuenta
                nuevo.num_cuenta_aux = k.num_cuenta
                nuevo.costo = int(0)
                nuevo.num_cuenta_iva = " "
                nuevo.valor_iva = int(0)
                
                if k.num_cuenta[0:4] == "2365":#esto es una retencion
                    nuevo.valor_contra = cpp_ser_pub.costo*(cpp_ser_pub.reten/100)
                else:#esto es el total de la CPP
                    nuevo.valor_contra = np.sum(vector_distri)-(cpp_ser_pub.costo*(cpp_ser_pub.reten/100))
                    
                nuevo.fecha = cpp_ser_pub.fecha
                
                nuevo.save()
                
                ii += 1     

            
            
    
    cpp_servi_detalle1 = cpp_servi_detalle.objects.filter(cpp_serv_public = id_)
    contexto={'cpp_servi_detalles':cpp_servi_detalle1}
    return render(request, 'servicios_publicos/cpp_servi_detalle_list.html',contexto)